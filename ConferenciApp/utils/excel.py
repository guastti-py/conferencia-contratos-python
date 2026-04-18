# ==========================================================================================
# UTILITÁRIO DE EXCEL
# Funções de formatação, colunas, nomes e checagens visuais no excel
# ==========================================================================================

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule


def ajustar_largura_colunas(ws):
    for col in ws.columns:
        tamanho_max = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = tamanho_max + 6


def aplicar_formatacao_padrao(ws, colunas_financeiras):
    fill_cinza = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    borda_fina = Side(style="thin", color="000000")
    borda_total = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")

    indices_financeiros = [
        cell.column for cell in ws[1] if cell.value in colunas_financeiras
    ]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento
            cell.border = borda_total

            if cell.row == 1:
                cell.fill = fill_cinza
                cell.font = fonte_cabecalho
            elif cell.column in indices_financeiros:
                cell.number_format = 'R$ #,##0.00'

    ajustar_largura_colunas(ws)

def obter_nome_arquivo_saida(nome_parceiro, codigo_parceiro, data_digitada, is_torra):
    data_formatada = data_digitada.replace("/", ".")
    codigo_seguro = codigo_parceiro.replace("/", "_")

    if is_torra:
        return f"Checagem - {nome_parceiro} {codigo_seguro} - {data_formatada}.xlsx"

    return f"Resumo {nome_parceiro} {codigo_seguro} - {data_formatada}.xlsx"    

def obter_indice_coluna_por_nome(ws, nome_coluna):
    nome_coluna = str(nome_coluna).strip().upper()

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if str(valor).strip().upper() == nome_coluna:
            return col

    return None

def encontrar_indice_coluna_por_nomes(ws, nomes_possiveis):
    for nome in nomes_possiveis:
        indice = obter_indice_coluna_por_nome(ws, nome)
        if indice is not None:
            return indice
    return None


def aplicar_formatacao_condicional_checagem(ws):
    verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    verde_font = Font(color="008000")

    vermelho_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    vermelho_font = Font(color="9C0006")

    for faixa in ["K4:K8", "K11:K15", "K18:K22"]:
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(operator="equal", formula=['"CONFERE"'], fill=verde_fill, font=verde_font)
        )
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(operator="equal", formula=['"NÃO CONFERE"'], fill=vermelho_fill, font=vermelho_font)
        )