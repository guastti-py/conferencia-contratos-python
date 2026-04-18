from tkinter import messagebox, filedialog
# ==========================================================================================
# PARCEIRO: PARCEIRO D
# Regras de conferência entre PDF do termo e Excel de resumo para os códigos 159, 199 e 200
# ==========================================================================================

import os
import pandas as pd
import pdfplumber

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from utils.arquivos import limpar_valor_monetario
from utils.excel import aplicar_formatacao_padrao


def extrair_dados_pdf_parceiro_d(caminhos_pdf):
    """
    Extrai do(s) PDF(s) do Termo Parceiro D:
    Colunas reais do PDF: col0=Nº CCB | col1=Emitente | col2=CPF | col3=Valor de Face | col4=Preço de Aquisição

    Retorna:
    - sets de CCBs e CPFs
    - soma total do Preço de Aquisição
    - quantidade total de contratos
    - lista de linhas completas para montar a aba no Excel
    """
    todos_ccbs = []
    todos_cpfs = []
    total_preco_aquisicao = 0.0
    linhas_aba = []  # [ccb, emitente, cpf, valor_face, preco_aquisicao]

    for caminho_pdf in caminhos_pdf:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                tabelas = pagina.extract_tables()

                for tabela in tabelas:
                    for linha in tabela:
                        if not linha or not linha[0]:
                            continue

                        col0 = str(linha[0]).strip().replace("\n", " ")

                        # Pula cabeçalho e linha de total
                        if "Nº da CCB" in col0 or col0.startswith("Total de") or col0 == "":
                            continue

                        # Layout real: col0=CCB, col1=Emitente, col2=CPF, col3=Val.Face, col4=Preço Aquisição
                        ccb   = col0
                        nome  = str(linha[1]).strip().replace("\n", " ") if len(linha) > 1 and linha[1] else ""
                        cpf   = str(linha[2]).strip().replace("\n", " ") if len(linha) > 2 and linha[2] else ""
                        vf    = limpar_valor_monetario(linha[3]) if len(linha) > 3 and linha[3] else 0.0
                        preco = limpar_valor_monetario(linha[4]) if len(linha) > 4 and linha[4] else 0.0

                        todos_ccbs.append(ccb)
                        todos_cpfs.append(cpf.replace(".", "").replace("-", "").strip())
                        total_preco_aquisicao += preco
                        linhas_aba.append([ccb, nome, cpf, vf, preco])

    return {
        "ccbs": set(todos_ccbs),
        "cpfs": set(todos_cpfs),
        "total_preco_aquisicao": round(total_preco_aquisicao, 2),
        "total_contratos": len(todos_ccbs),
        "linhas_aba": linhas_aba
    }


def salvar_aba_dados_pdf_parceiro_d(wb, linhas_aba):
    """Cria (ou recria) a aba 'Dados PDF' com as linhas extraídas do(s) Termo(s) Parceiro D."""
    nome_aba = "Dados PDF"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws_pdf = wb.create_sheet(title=nome_aba)

    cabecalho = ["Nº da CCB", "Emitente", "CPF do Emitente", "Valor de Face", "Preço de Aquisição"]
    ws_pdf.append(cabecalho)

    for linha in linhas_aba:
        ws_pdf.append(linha)

    # Linha de total
    ultima_linha = ws_pdf.max_row + 1
    ws_pdf.cell(row=ultima_linha, column=1, value="TOTAL")
    total_vf    = sum(l[3] for l in linhas_aba)
    total_preco = sum(l[4] for l in linhas_aba)
    ws_pdf.cell(row=ultima_linha, column=4, value=round(total_vf, 2))
    ws_pdf.cell(row=ultima_linha, column=5, value=round(total_preco, 2))

    aplicar_formatacao_padrao(ws_pdf, ["Valor de Face", "Preço de Aquisição"])

    # Destaca linha de total
    fill_total = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fonte_total = Font(bold=True)
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alin = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        c = ws_pdf.cell(row=ultima_linha, column=col)
        c.fill = fill_total
        c.font = fonte_total
        c.border = borda
        c.alignment = alin
        if col in (4, 5):
            c.number_format = 'R$ #,##0.00'


def escrever_resumo_termo_parceiro_d(ws, linha_inicio, col_inicio, dados_pdf):
    """Escreve o bloco 'Resumo Termo' (azul) com os dados extraídos do PDF."""
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")
    alinhamento = Alignment(horizontal="center", vertical="center")

    fonte_titulo = Font(name="Aptos Narrow", size=13, bold=True, color="FFFFFF")
    fonte_valor = Font(name="Aptos Narrow", size=12, bold=True)

    cor_azul = "2E75B6"

    # Cabeçalho "Resumo Termo"
    ws.merge_cells(
        start_row=linha_inicio, start_column=col_inicio,
        end_row=linha_inicio, end_column=col_inicio + 1
    )
    cel_titulo = ws.cell(row=linha_inicio, column=col_inicio, value="Resumo Termo")
    cel_titulo.font = fonte_titulo
    cel_titulo.fill = PatternFill(start_color=cor_azul, end_color=cor_azul, fill_type="solid")
    cel_titulo.alignment = alinhamento
    ws.cell(row=linha_inicio, column=col_inicio).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa
    )
    ws.cell(row=linha_inicio, column=col_inicio + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa
    )

    # Op.:
    cel_r1 = ws.cell(row=linha_inicio + 1, column=col_inicio, value="Op.:")
    cel_r1.font = fonte_valor
    cel_r1.alignment = alinhamento
    cel_r1.border = Border(left=borda_grossa, right=borda_fina, top=borda_fina, bottom=borda_fina)

    cel_v1 = ws.cell(row=linha_inicio + 1, column=col_inicio + 1, value=dados_pdf["total_contratos"])
    cel_v1.font = fonte_valor
    cel_v1.alignment = alinhamento
    cel_v1.border = Border(left=borda_fina, right=borda_grossa, top=borda_fina, bottom=borda_fina)

    # V.Cessão:
    cel_r2 = ws.cell(row=linha_inicio + 2, column=col_inicio, value="V.Cessão:")
    cel_r2.font = fonte_valor
    cel_r2.alignment = alinhamento
    cel_r2.border = Border(left=borda_grossa, right=borda_fina, top=borda_fina, bottom=borda_grossa)

    cel_v2 = ws.cell(row=linha_inicio + 2, column=col_inicio + 1, value=dados_pdf["total_preco_aquisicao"])
    cel_v2.font = fonte_valor
    cel_v2.alignment = alinhamento
    cel_v2.number_format = 'R$ #,##0.00'
    cel_v2.border = Border(left=borda_fina, right=borda_grossa, top=borda_fina, bottom=borda_grossa)


def escrever_checagem_parceiro_d(ws, linha_inicio, col_inicio,
                             linha_resumo_sistema, col_resumo_sistema,
                             linha_resumo_termo, col_resumo_termo,
                             dados_resumo_sistema, dados_pdf):
    """
    Escreve a tabela CHECAGEM com 4 itens: CPF, CCB, Nº Op., CESSÃO
    Compara dados do Resumo Sistema (Excel) com dados do PDF.
    """
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")
    alinhamento = Alignment(horizontal="center", vertical="center")

    fonte_titulo = Font(name="Aptos Narrow", size=13, bold=True, color="000000")
    fonte_header = Font(name="Aptos Narrow", size=11, bold=True)
    fonte_item = Font(name="Aptos Narrow", size=11)

    cor_amarelo = "FFFF00"

    verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    verde_font = Font(color="008000", bold=True)
    vermelho_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    vermelho_font = Font(color="9C0006", bold=True)

    # Cabeçalho CHECAGEM
    ws.merge_cells(
        start_row=linha_inicio, start_column=col_inicio,
        end_row=linha_inicio, end_column=col_inicio + 1
    )
    cel_check = ws.cell(row=linha_inicio, column=col_inicio, value="CHECAGEM")
    cel_check.font = fonte_titulo
    cel_check.fill = PatternFill(start_color=cor_amarelo, end_color=cor_amarelo, fill_type="solid")
    cel_check.alignment = alinhamento
    ws.cell(row=linha_inicio, column=col_inicio).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa
    )
    ws.cell(row=linha_inicio, column=col_inicio + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa
    )

    # Sub-cabeçalhos: PRODUTO | STATUS
    cel_prod = ws.cell(row=linha_inicio + 1, column=col_inicio, value="PRODUTO")
    cel_prod.font = fonte_header
    cel_prod.alignment = alinhamento
    cel_prod.border = Border(left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa)

    cel_stat = ws.cell(row=linha_inicio + 1, column=col_inicio + 1, value="STATUS")
    cel_stat.font = fonte_header
    cel_stat.alignment = alinhamento
    cel_stat.border = Border(left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa)

    # ---- Lógica de comparação ----
    cpfs_excel = set(dados_resumo_sistema.get("cpfs", []))
    ccbs_excel = set(dados_resumo_sistema.get("ccbs", []))
    nop_excel = dados_resumo_sistema.get("total_contratos", 0)
    cessao_excel = dados_resumo_sistema.get("total_cessao", 0.0)

    cpfs_pdf = dados_pdf["cpfs"]
    ccbs_pdf = dados_pdf["ccbs"]
    nop_pdf = dados_pdf["total_contratos"]
    cessao_pdf = dados_pdf["total_preco_aquisicao"]

    cpf_ok = (cpfs_excel == cpfs_pdf)
    ccb_ok = (ccbs_excel == ccbs_pdf)
    nop_ok = (nop_excel == nop_pdf)
    cessao_ok = (round(cessao_excel, 2) == round(cessao_pdf, 2))

    itens = [
        ("CPF", cpf_ok),
        ("CCB", ccb_ok),
        ("Nº Op.", nop_ok),
        ("CESSÃO", cessao_ok),
    ]

    for i, (produto, ok) in enumerate(itens):
        linha = linha_inicio + 2 + i
        eh_ultima = i == len(itens) - 1
        status = "CONFERE" if ok else "NÃO CONFERE"

        cel_p = ws.cell(row=linha, column=col_inicio, value=produto)
        cel_p.font = fonte_item
        cel_p.alignment = alinhamento
        cel_p.border = Border(
            left=borda_grossa, right=borda_fina,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina
        )

        cel_s = ws.cell(row=linha, column=col_inicio + 1, value=status)
        cel_s.alignment = alinhamento
        cel_s.border = Border(
            left=borda_fina, right=borda_grossa,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina
        )
        if ok:
            cel_s.fill = verde_fill
            cel_s.font = verde_font
        else:
            cel_s.fill = vermelho_fill
            cel_s.font = vermelho_font


def obter_dados_resumo_sistema_parceiro_d(caminho_excel):
    """
    Lê a aba 'Dados Originais' do arquivo de resumo Parceiro D e retorna
    os sets de CPFs, CCBs, total de contratos e V.Cessão total.
    """
    df = pd.read_excel(caminho_excel, sheet_name="Dados Originais")
    df.columns = df.columns.str.strip()

    # Coluna de contrato/CCB
    col_ccb = None
    for col in df.columns:
        if "CONTRATO" in col.upper() or "CCB" in col.upper():
            col_ccb = col
            break

    # Coluna CPF
    col_cpf = None
    for col in df.columns:
        if "CPF" in col.upper():
            col_cpf = col
            break

    # Coluna VALOR_CESSAO
    col_cessao = None
    for col in df.columns:
        if "CESSAO" in col.upper() or "CESSÃO" in col.upper():
            col_cessao = col
            break

    ccbs = set()
    if col_ccb:
        ccbs = set(df[col_ccb].astype(str).str.strip().tolist())

    cpfs = set()
    if col_cpf:
        cpfs = set(df[col_cpf].astype(str).str.strip().str.replace(".", "", regex=False).str.replace("-", "", regex=False).tolist())

    cessao_total = 0.0
    if col_cessao:
        cessao_total = pd.to_numeric(df[col_cessao], errors="coerce").fillna(0).sum()

    return {
        "ccbs": ccbs,
        "cpfs": cpfs,
        "total_contratos": len(df),
        "total_cessao": round(cessao_total, 2)
    }


def fazer_conferencia_parceiro_d():
    """
    Fluxo de conferência para parceiros Parceiro D (159, 199, 200):
    1. Seleciona o Excel de Resumo já gerado
    2. Seleciona um ou mais PDFs do Termo
    3. Extrai dados do(s) PDF(s)
    4. Escreve 'Resumo Termo' e tabela 'CHECAGEM' na aba Resumo
    5. Salva o arquivo
    """
    messagebox.showinfo(
        "Conferência Parceiro D",
        "1º Selecione o arquivo EXCEL de Resumo (já gerado)\n"
        "2º Selecione o(s) PDF(s) do Termo de Endosso"
    )

    caminho_excel = filedialog.askopenfilename(
        title="Selecionar Resumo Excel - Parceiro D",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_excel:
        return

    # Coleta de PDFs — pergunta se tem mais de um
    caminhos_pdf = []
    while True:
        caminho_pdf = filedialog.askopenfilename(
            title=f"Selecionar PDF do Termo ({len(caminhos_pdf) + 1}º)",
            filetypes=[("PDF", "*.pdf")]
        )
        if not caminho_pdf:
            if not caminhos_pdf:
                messagebox.showwarning("Atenção", "Nenhum PDF selecionado. Operação cancelada.")
                return
            break
        caminhos_pdf.append(caminho_pdf)

        resposta = messagebox.askyesno(
            "Mais PDFs?",
            f"{len(caminhos_pdf)} PDF(s) selecionado(s).\n\nDeseja adicionar mais um PDF?"
        )
        if not resposta:
            break

    try:
        # Extrai dados do(s) PDF(s)
        dados_pdf = extrair_dados_pdf_parceiro_d(caminhos_pdf)

        # Lê dados do resumo sistema (Excel)
        dados_sistema = obter_dados_resumo_sistema_parceiro_d(caminho_excel)

        # Abre o workbook
        wb = load_workbook(caminho_excel)

        if "Resumo" not in wb.sheetnames:
            raise Exception("A aba 'Resumo' não foi encontrada no Excel.")

        # Cria aba com os dados brutos do(s) PDF(s)
        salvar_aba_dados_pdf_parceiro_d(wb, dados_pdf["linhas_aba"])

        ws = wb["Resumo"]

        # Encontra a última linha usada no bloco Resumo Sistema (cols B e C)
        ultima_linha_usada = 1
        for row in ws.iter_rows(min_col=2, max_col=3):
            for cell in row:
                if cell.value is not None:
                    ultima_linha_usada = max(ultima_linha_usada, cell.row)

        linha_resumo_termo = ultima_linha_usada + 2  # 1 linha de espaço
        linha_checagem = 2  # Começa na linha 2, coluna E/F (col 5)

        # Ajusta larguras das colunas para os novos blocos
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 18

        # Escreve Resumo Termo (bloco azul) — abaixo do Resumo Sistema, na mesma coluna B/C
        escrever_resumo_termo_parceiro_d(ws, linha_resumo_termo, 2, dados_pdf)

        # Escreve CHECAGEM — ao lado, nas colunas E/F
        escrever_checagem_parceiro_d(
            ws,
            linha_inicio=linha_checagem,
            col_inicio=5,  # coluna E
            linha_resumo_sistema=2,
            col_resumo_sistema=2,
            linha_resumo_termo=linha_resumo_termo,
            col_resumo_termo=2,
            dados_resumo_sistema=dados_sistema,
            dados_pdf=dados_pdf
        )

        try:
            wb.save(caminho_excel)
        except PermissionError:
            messagebox.showerror(
                "Arquivo aberto",
                f"Não foi possível salvar o arquivo pois ele está aberto no Excel.\n\n"
                f"Feche o arquivo e tente novamente:\n{caminho_excel}"
            )
            return

        messagebox.showinfo(
            "Sucesso",
            f"Conferência Parceiro D concluída!\n"
            f"{len(caminhos_pdf)} PDF(s) processado(s).\n"
            f"Total contratos PDF: {dados_pdf['total_contratos']}\n"
            f"V.Cessão PDF: R$ {dados_pdf['total_preco_aquisicao']:,.2f}"
        )
        os.startfile(caminho_excel)

    except Exception as erro:
        messagebox.showerror("Erro na Conferência Parceiro D", str(erro))