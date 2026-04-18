import os
import re
from tkinter import filedialog, messagebox
import tkinter as tk

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ======================================================================================
# UTILITÁRIOS
# ======================================================================================

def carregar_csv_generico(caminho_arquivo):
    try:
        return pd.read_csv(caminho_arquivo, sep=None, engine="python", encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(caminho_arquivo, sep=None, engine="python", encoding="latin-1")


def padronizar_colunas(df):
    df.columns = df.columns.astype(str).str.strip().str.upper()
    return df


def converter_colunas_numericas(df):
    colunas_monetarias = ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF"]

    for col in colunas_monetarias:
        if col not in df.columns:
            df[col] = 0
            continue

        df[col] = (
            df[col]
            .astype(str)
            .str.replace("R$", "", regex=False)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
            .str.strip()
        )
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def ajustar_largura_colunas(ws):
    for col in ws.columns:
        tamanho_max = 0
        letra = get_column_letter(col[0].column)
        for cell in col:
            valor = "" if cell.value is None else str(cell.value)
            if len(valor) > tamanho_max:
                tamanho_max = len(valor)
        ws.column_dimensions[letra].width = tamanho_max + 4


def aplicar_formatacao_padrao(ws, colunas_financeiras=None):
    if colunas_financeiras is None:
        colunas_financeiras = []

    fill_cabecalho = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    indices_financeiros = []
    for cell in ws[1]:
        if str(cell.value).strip().upper() in [c.upper() for c in colunas_financeiras]:
            indices_financeiros.append(cell.column)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento
            cell.border = borda

            if cell.row == 1:
                cell.fill = fill_cabecalho
                cell.font = fonte_cabecalho
            elif cell.column in indices_financeiros:
                cell.number_format = 'R$ #,##0.00'

    ajustar_largura_colunas(ws)


def obter_indice_coluna_por_nome(ws, nome_coluna):
    nome_coluna = str(nome_coluna).strip().upper()
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if str(valor).strip().upper() == nome_coluna:
            return col
    return None


def mover_resumo_para_primeira_aba(wb):
    if "Resumo" in wb.sheetnames:
        aba_resumo = wb["Resumo"]
        wb._sheets.remove(aba_resumo)
        wb._sheets.insert(0, aba_resumo)


# ======================================================================================
# COLUNAS ADICIONAIS NAS OPERAÇÕES
# ======================================================================================

def adicionar_colunas_calculadas(ws):
    headers = [str(ws.cell(row=1, column=col).value).strip().upper() for col in range(1, ws.max_column + 1)]

    def get_col(nome):
        return headers.index(nome) + 1 if nome in headers else None

    col_vb = get_col("VALOR_BRUTO")
    col_vl = get_col("VALOR_LIQUIDO")
    col_iof = get_col("IOF")
    col_cc = get_col("CONTA_CORRENTE")

    if not all([col_vb, col_vl, col_iof, col_cc]):
        raise Exception("Colunas necessárias não encontradas: CONTA_CORRENTE, VALOR_BRUTO, VALOR_LIQUIDO e IOF.")

    insert_pos = col_cc + 1

    novas_colunas = [
        "VALOR DO SEGURO",
        "ÁGIO",
        "PAGAMENTO EFETUADO",
        "DATA DO PAGAMENTO",
        "MOTIVO DEV."
    ]

    for i, nome in enumerate(novas_colunas):
        ws.insert_cols(insert_pos + i)
        ws.cell(row=1, column=insert_pos + i, value=nome)

    col_seguro = insert_pos
    col_agio = insert_pos + 1

    letra_vb = get_column_letter(col_vb)
    letra_vl = get_column_letter(col_vl)
    letra_iof = get_column_letter(col_iof)

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_seguro).value = f"={letra_vb}{row}-{letra_vl}{row}-{letra_iof}{row}"
        ws.cell(row=row, column=col_agio).value = f"={letra_vb}{row}*0.006"

    return {
        "VB": col_vb,
        "VL": col_vl,
        "IOF": col_iof,
        "SEGURO": col_seguro,
        "AGIO": col_agio
    }


# ======================================================================================
# RESUMO OPERAÇÕES
# ======================================================================================

def criar_resumo_operacoes(wb):
    if "Resumo" in wb.sheetnames:
        ws = wb["Resumo"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Resumo")

    ws_oper = wb["Operações"]

    col_vb = obter_indice_coluna_por_nome(ws_oper, "VALOR_BRUTO")
    col_vl = obter_indice_coluna_por_nome(ws_oper, "VALOR_LIQUIDO")
    col_iof = obter_indice_coluna_por_nome(ws_oper, "IOF")

    if not all([col_vb, col_vl, col_iof]):
        raise Exception("Não encontrei as colunas VALOR_BRUTO, VALOR_LIQUIDO e IOF nos CSVs da PARCEIRO Z.")

    letra_vb = get_column_letter(col_vb)
    letra_vl = get_column_letter(col_vl)
    letra_iof = get_column_letter(col_iof)

    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(size=14, bold=True, color="000000")
    fonte_texto = Font(size=12, bold=True, color="000000")
    fill_titulo = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws.merge_cells("A1:B1")
    ws["A1"] = "Resumo operações"
    ws["A1"].font = fonte_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = alinhamento
    ws["A1"].border = borda
    ws["B1"].border = borda

    itens = [
        ("Op.:", '=COUNTA(Operações!A:A)-1'),
        ("V.B.:", f'=SUM(Operações!{letra_vb}:{letra_vb})'),
        ("V.L.:", f'=SUM(Operações!{letra_vl}:{letra_vl})'),
        ("IOF:", f'=SUM(Operações!{letra_iof}:{letra_iof})'),
        ("ÁGIO:", '=B3*0.006'),
        ("V. ENDOSSO:", '=B3+B6'),
        ("V. SEGURO:", '=B3-B4-B5'),
    ]

    linha = 2
    for rotulo, formula in itens:
        ws.cell(row=linha, column=1, value=rotulo)
        ws.cell(row=linha, column=2, value=formula)

        ws.cell(row=linha, column=1).font = fonte_texto
        ws.cell(row=linha, column=2).font = fonte_texto
        ws.cell(row=linha, column=1).alignment = alinhamento
        ws.cell(row=linha, column=2).alignment = alinhamento
        ws.cell(row=linha, column=1).border = borda
        ws.cell(row=linha, column=2).border = borda

        if rotulo != "Op.:":
            ws.cell(row=linha, column=2).number_format = 'R$ #,##0.00'

        linha += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    # --- Bloco TERMO (coluna D) ---
    ws.merge_cells("D1:E1")
    ws["D1"] = "TERMO"
    ws["D1"].font = fonte_titulo
    ws["D1"].fill = fill_titulo
    ws["D1"].alignment = alinhamento
    ws["D1"].border = borda
    ws["E1"].border = borda

    itens_termo = [
        ("nº op", '=COUNTA(Operações!A:A)-1'),
        ("V.CONF", '=E3+E6'),  # V.B + ÁGIO  →  B3=VB, B6=ÁGIO
    ]

    # Preenchemos D2:E3 (2 itens), deixando D4–D7 vazios
    for i, (rotulo, formula) in enumerate(itens_termo):
        row = 2 + i
        ws.cell(row=row, column=4, value=rotulo)
        ws.cell(row=row, column=5, value=formula)

        ws.cell(row=row, column=4).font = fonte_texto
        ws.cell(row=row, column=5).font = fonte_texto
        ws.cell(row=row, column=4).alignment = alinhamento
        ws.cell(row=row, column=5).alignment = alinhamento
        ws.cell(row=row, column=4).border = borda
        ws.cell(row=row, column=5).border = borda

        if rotulo != "nº op":
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'

    # Célula explícita D8 = "V.CONF", E8 = V.B + ÁGIO (igual ao item acima, fixado na linha 8)
    ws.cell(row=8, column=4, value="V.CONF")
    ws.cell(row=8, column=5, value="=E3+E6")
    ws.cell(row=8, column=4).font = fonte_texto
    ws.cell(row=8, column=5).font = fonte_texto
    ws.cell(row=8, column=4).alignment = alinhamento
    ws.cell(row=8, column=5).alignment = alinhamento
    ws.cell(row=8, column=4).border = borda
    ws.cell(row=8, column=5).border = borda
    ws.cell(row=8, column=5).number_format = 'R$ #,##0.00'

    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22


# ======================================================================================
# RESUMO TESOURARIA
# ======================================================================================

def criar_resumo_pago_no_mesmo_resumo(wb):
    ws_resumo = wb["Resumo"]
    ws_tes = wb["OP Tesouraria"]

    col_pagto = obter_indice_coluna_por_nome(ws_tes, "PAGAMENTO EFETUADO")
    col_vb = obter_indice_coluna_por_nome(ws_tes, "VALOR_BRUTO")
    col_vl = obter_indice_coluna_por_nome(ws_tes, "VALOR_LIQUIDO")
    col_iof = obter_indice_coluna_por_nome(ws_tes, "IOF")

    if not all([col_pagto, col_vb, col_vl, col_iof]):
        raise Exception("Não encontrei as colunas PAGAMENTO EFETUADO, VALOR_BRUTO, VALOR_LIQUIDO e IOF na planilha da tesouraria.")

    letra_pagto = get_column_letter(col_pagto)
    letra_vb = get_column_letter(col_vb)
    letra_vl = get_column_letter(col_vl)
    letra_iof = get_column_letter(col_iof)

    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(size=14, bold=True, color="000000")
    fonte_texto = Font(size=12, bold=True, color="000000")
    fill_titulo = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws_resumo.merge_cells("D1:E1")
    ws_resumo["D1"] = "Resumo pago"
    ws_resumo["D1"].font = fonte_titulo
    ws_resumo["D1"].fill = fill_titulo
    ws_resumo["D1"].alignment = alinhamento
    ws_resumo["D1"].border = borda
    ws_resumo["E1"].border = borda

    itens = [
        ("Op.:", f'=COUNTIF(\'OP Tesouraria\'!{letra_pagto}:{letra_pagto},"OK")'),
        ("V.B.:", f'=SUMIF(\'OP Tesouraria\'!{letra_pagto}:{letra_pagto},"OK",\'OP Tesouraria\'!{letra_vb}:{letra_vb})'),
        ("V.L.:", f'=SUMIF(\'OP Tesouraria\'!{letra_pagto}:{letra_pagto},"OK",\'OP Tesouraria\'!{letra_vl}:{letra_vl})'),
        ("IOF:", f'=SUMIF(\'OP Tesouraria\'!{letra_pagto}:{letra_pagto},"OK",\'OP Tesouraria\'!{letra_iof}:{letra_iof})'),
        ("ÁGIO:", '=E3*0.006'),
        ("V. SEGURO:", '=E3-E4-E5'),
    ]

    linha = 2
    for rotulo, formula in itens:
        ws_resumo.cell(row=linha, column=4, value=rotulo)
        ws_resumo.cell(row=linha, column=5, value=formula)

        ws_resumo.cell(row=linha, column=4).font = fonte_texto
        ws_resumo.cell(row=linha, column=5).font = fonte_texto
        ws_resumo.cell(row=linha, column=4).alignment = alinhamento
        ws_resumo.cell(row=linha, column=5).alignment = alinhamento
        ws_resumo.cell(row=linha, column=4).border = borda
        ws_resumo.cell(row=linha, column=5).border = borda

        if rotulo != "Op.:":
            ws_resumo.cell(row=linha, column=5).number_format = 'R$ #,##0.00'

        linha += 1

    ws_resumo.column_dimensions["D"].width = 18
    ws_resumo.column_dimensions["E"].width = 22


# ======================================================================================
# ETAPA 1 — GERAR RESUMO + SEGUNDO ARQUIVO
# ======================================================================================

def processar_parceiro_c(root, entry_data, progress_bar, event=None):
    data_digitada = entry_data.get().strip()

    if not data_digitada:
        messagebox.showerror("Erro", "Digite a data.")
        return

    caminho_288 = filedialog.askopenfilename(
        title="Selecione o CSV PARCEIRO Z 288",
        filetypes=[("CSV", "*.csv")]
    )
    if not caminho_288:
        return

    caminho_296 = filedialog.askopenfilename(
        title="Selecione o CSV PARCEIRO Z 296",
        filetypes=[("CSV", "*.csv")]
    )
    if not caminho_296:
        return

    try:
        progress_bar.pack(pady=10)
        progress_bar.set(0.20)
        root.update_idletasks()

        df_288 = carregar_csv_generico(caminho_288)
        df_296 = carregar_csv_generico(caminho_296)

        df_288 = padronizar_colunas(df_288)
        df_296 = padronizar_colunas(df_296)

        df_288 = converter_colunas_numericas(df_288)
        df_296 = converter_colunas_numericas(df_296)

        df_total = pd.concat([df_288, df_296], ignore_index=True)

        progress_bar.set(0.50)
        root.update_idletasks()

        pasta_saida = os.path.dirname(caminho_288)

        nome_arquivo_resumo = f"Checagem - PARCEIRO Z - {data_digitada}.xlsx"
        caminho_saida_resumo = os.path.join(pasta_saida, nome_arquivo_resumo)

        nome_arquivo_operacoes = f"PARCEIRO Z - {data_digitada}.xlsx"
        caminho_saida_operacoes = os.path.join(pasta_saida, nome_arquivo_operacoes)

        # ---------------------------------
        # ARQUIVO 1: CHECAGEM - PARCEIRO Z
        # ---------------------------------
        wb = Workbook()
        ws_oper = wb.active
        ws_oper.title = "Operações"

        for linha in dataframe_to_rows(df_total, index=False, header=True):
            ws_oper.append(linha)

        adicionar_colunas_calculadas(ws_oper)
        aplicar_formatacao_padrao(
            ws_oper,
            ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "VALOR DO SEGURO", "ÁGIO"]
        )

        criar_resumo_operacoes(wb)
        mover_resumo_para_primeira_aba(wb)
        wb.save(caminho_saida_resumo)

        # ---------------------------------
        # ARQUIVO 2: PARCEIRO Z
        # ---------------------------------
        wb_operacoes = Workbook()
        ws_operacoes = wb_operacoes.active
        ws_operacoes.title = "Operações"

        for linha in dataframe_to_rows(df_total, index=False, header=True):
            ws_operacoes.append(linha)

        indices = adicionar_colunas_calculadas(ws_operacoes)
        aplicar_formatacao_padrao(
            ws_operacoes,
            ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "VALOR DO SEGURO", "ÁGIO"]
        )

        ultima_linha = ws_operacoes.max_row + 1
        ws_operacoes.cell(row=ultima_linha, column=1, value="TOTAL")

        for chave in ["VB", "VL", "IOF", "SEGURO", "AGIO"]:
            col = indices[chave]
            letra = get_column_letter(col)
            ws_operacoes.cell(
                row=ultima_linha,
                column=col,
                value=f"=SUM({letra}2:{letra}{ultima_linha - 1})"
            )
            ws_operacoes.cell(row=ultima_linha, column=col).number_format = 'R$ #,##0.00'

        wb_operacoes.save(caminho_saida_operacoes)

        progress_bar.set(1.0)
        root.update_idletasks()
        progress_bar.pack_forget()

        os.startfile(pasta_saida)
        messagebox.showinfo(
            "Sucesso",
            f"Arquivos gerados com sucesso:\n\n{nome_arquivo_resumo}\n{nome_arquivo_operacoes}"
        )

    except Exception as e:
        progress_bar.pack_forget()
        messagebox.showerror("Erro", str(e))


# ======================================================================================
# ETAPA 2 — CONFERÊNCIA TESOURARIA
# ======================================================================================

def conferencia_parceiro_c(root, entry_data, progress_bar, event=None):
    caminho_resumo = filedialog.askopenfilename(
        title="Selecione o arquivo de resumo PARCEIRO Z",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_resumo:
        return

    caminho_tesouraria = filedialog.askopenfilename(
        title="Selecione o arquivo Excel da tesouraria PARCEIRO Z",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_tesouraria:
        return

    try:
        progress_bar.pack(pady=10)
        progress_bar.set(0.30)
        root.update_idletasks()

        df_tes = pd.read_excel(caminho_tesouraria)
        df_tes = padronizar_colunas(df_tes)

        for col in ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF"]:
            if col in df_tes.columns:
               df_tes[col] = pd.to_numeric(df_tes[col], errors="coerce").fillna(0)

        wb = load_workbook(caminho_resumo)

        if "OP Tesouraria" in wb.sheetnames:
            del wb["OP Tesouraria"]

        ws_tes = wb.create_sheet("OP Tesouraria")

        for linha in dataframe_to_rows(df_tes, index=False, header=True):
            ws_tes.append(linha)

        aplicar_formatacao_padrao(
            ws_tes,
            ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF"]
        )

        progress_bar.set(0.70)
        root.update_idletasks()

        if "Resumo" not in wb.sheetnames:
            raise Exception("A aba 'Resumo' não foi encontrada no arquivo de checagem.")

        criar_resumo_pago_no_mesmo_resumo(wb)
        mover_resumo_para_primeira_aba(wb)

        wb.save(caminho_resumo)

        progress_bar.set(1.0)
        root.update_idletasks()
        progress_bar.pack_forget()

        os.startfile(caminho_resumo)
        messagebox.showinfo("Sucesso", "Conferência Tesouraria PARCEIRO Z incluída com sucesso!")

    except Exception as e:
        progress_bar.pack_forget()
        messagebox.showerror("Erro", str(e))

# ======================================================================================
# ETAPA 3 — CONFERÊNCIA PDF (ENDOSSO)
# ======================================================================================

def _extrair_dados_pdf_endosso(caminho_pdf):
    """
    Extrai dados do PDF do Termo de Endosso.
    Retorna lista de dicts com: ccb, emitente, cpf_cnpj, valor_principal, preco_endosso
    """
    registros = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            tabelas = pagina.extract_tables()
            for tabela in tabelas:
                for linha in tabela:
                    if linha is None:
                        continue
                    # Filtrar linhas com dados de CCB (começam com 'A' seguido de números)
                    ccb = str(linha[0]).strip() if linha[0] else ""
                    if not (ccb.startswith("A") and len(ccb) > 5):
                        continue
                    try:
                        emitente = str(linha[1]).strip() if len(linha) > 1 and linha[1] else ""
                        cpf = str(linha[2]).strip() if len(linha) > 2 and linha[2] else ""
                        val_principal_raw = str(linha[3]).strip() if len(linha) > 3 and linha[3] else "0"
                        preco_endosso_raw = str(linha[4]).strip() if len(linha) > 4 and linha[4] else "0"

                        def parse_valor(s):
                            s = s.replace("R$", "").replace("\n", "").replace(" ", "")
                            s = s.replace(".", "").replace(",", ".")
                            return float(s) if s else 0.0

                        registros.append({
                            "Nº da CCB": ccb,
                            "Emitente": emitente,
                            "CNPJ/CPF do Emitente": cpf,
                            "Valor de Principal": parse_valor(val_principal_raw),
                            "Preço de Endosso da CCB": parse_valor(preco_endosso_raw),
                        })
                    except Exception:
                        continue
    return registros


def _aplicar_formatacao_tabela_pdf(ws):
    """Formata a aba de dados do PDF."""
    fill_cab = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fonte_cab = Font(color="FFFFFF", bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    fmt_money = 'R$ #,##0.00'

    colunas_money = {"Valor de Principal", "Preço de Endosso da CCB"}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento
            cell.border = borda
            if cell.row == 1:
                cell.fill = fill_cab
                cell.font = fonte_cab
            else:
                header = ws.cell(row=1, column=cell.column).value
                if header in colunas_money:
                    cell.number_format = fmt_money

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


def _gravar_dados_pdf_na_aba(ws, registros, primeira_vez):
    """
    Se primeira_vez=True, escreve cabeçalho + dados.
    Se primeira_vez=False, apenas acrescenta dados a partir da última linha preenchida.
    """
    colunas = ["Nº da CCB", "Emitente", "CNPJ/CPF do Emitente", "Valor de Principal", "Preço de Endosso da CCB"]

    if primeira_vez:
        ws.append(colunas)
        for reg in registros:
            ws.append([reg[c] for c in colunas])
    else:
        # Acha última linha com dados
        ultima = ws.max_row
        for reg in registros:
            ultima += 1
            for j, c in enumerate(colunas, start=1):
                ws.cell(row=ultima, column=j, value=reg[c])


def _normalizar_ccb(valor):
    """
    Normaliza o contrato/CCB para comparação:
    - remove espaços
    - remove pontuação
    - deixa maiúsculo
    """
    if valor is None:
        return ""
    valor = str(valor).strip().upper()
    valor = "".join(ch for ch in valor if ch.isalnum())
    return valor

def _normalizar_contrato(valor):
    """
    Normaliza contrato/CCB para comparação:
    - converte para texto
    - remove espaços
    - remove .0 de números vindos do Excel
    - mantém apenas dígitos
    """
    if valor is None:
        return ""

    valor = str(valor).strip()

    if valor.endswith(".0"):
        valor = valor[:-2]

    # mantém só números
    valor = "".join(ch for ch in valor if ch.isdigit())

    return valor

def _normalizar_contrato_parceiro_c(valor):
    """
    Normaliza contrato/CCB da PARCEIRO Z para comparação.

    Exemplos:
    A7730858-000 -> A7730858
    A7730858     -> A7730858
    a7730858/000 -> A7730858
    7730858      -> 7730858
    """
    if valor is None:
        return ""

    texto = str(valor).strip().upper()

    if not texto:
        return ""

    # remove .0 vindo do Excel
    if texto.endswith(".0"):
        texto = texto[:-2]

    # remove espaços
    texto = texto.replace(" ", "")

    # tira tudo depois de hífen ou barra
    texto = texto.split("-")[0]
    texto = texto.split("/")[0]

    # mantém só letras e números
    texto = "".join(ch for ch in texto if ch.isalnum())

    return texto


def _calcular_diferenca_ccbs(wb):
    """
    Compara:
    - aba OP Tesouraria -> coluna CONTRATO
    - aba Dados PDF     -> coluna Nº da CCB

    Retorna os contratos que estão na OP Tesouraria
    mas NÃO estão no PDF.
    """
    if "Dados PDF" not in wb.sheetnames or "OP Tesouraria" not in wb.sheetnames:
        return []

    ws_pdf = wb["Dados PDF"]
    ws_tes = wb["OP Tesouraria"]

    # localizar coluna CONTRATO na OP Tesouraria
    col_tes_contrato = None
    for col in range(1, ws_tes.max_column + 1):
        cab = str(ws_tes.cell(row=1, column=col).value or "").strip().upper()
        if cab == "CONTRATO":
            col_tes_contrato = col
            break

    if col_tes_contrato is None:
        raise Exception("Não encontrei a coluna 'CONTRATO' na aba OP Tesouraria.")

    # localizar coluna Nº da CCB no PDF Endosso
    col_pdf_ccb = None
    for col in range(1, ws_pdf.max_column + 1):
        cab = str(ws_pdf.cell(row=1, column=col).value or "").strip().upper()
        if "CCB" in cab:
            col_pdf_ccb = col
            break

    if col_pdf_ccb is None:
        raise Exception("Não encontrei a coluna 'Nº da CCB' na aba PDF Endosso.")

    # contratos do PDF normalizados
    contratos_pdf = set()
    for row in range(2, ws_pdf.max_row + 1):
        valor_pdf = ws_pdf.cell(row=row, column=col_pdf_ccb).value
        contrato_pdf = _normalizar_contrato_parceiro_c(valor_pdf)
        if contrato_pdf:
            contratos_pdf.add(contrato_pdf)

    # cabeçalhos completos da tesouraria para montar a aba final
    cabecalhos_tes = [
        str(ws_tes.cell(row=1, column=col).value or "").strip()
        for col in range(1, ws_tes.max_column + 1)
    ]

    # contratos da tesouraria que não existem no PDF
    faltando_no_pdf = []

    for row in range(2, ws_tes.max_row + 1):
        valor_tes = ws_tes.cell(row=row, column=col_tes_contrato).value
        contrato_tes = _normalizar_contrato_parceiro_c(valor_tes)

        if contrato_tes and contrato_tes not in contratos_pdf:
            linha_dict = {
                cabecalhos_tes[col - 1]: ws_tes.cell(row=row, column=col).value
                for col in range(1, ws_tes.max_column + 1)
            }

            faltando_no_pdf.append({
                "CONTRATO": valor_tes,
                "linha_completa": linha_dict
            })

    return faltando_no_pdf


def _criar_aba_contratos_faltando(wb, contratos_faltando):
    """Cria ou recria a aba com os contratos da OP Tesouraria que não estão no PDF."""
    nome_aba = "Contratos Faltantes PDF"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws = wb.create_sheet(nome_aba)

    if not contratos_faltando:
        ws.append(["CONTRATO"])
        ws.append(["Nenhum contrato faltante encontrado."])
        _aplicar_formatacao_tabela_pdf(ws)
        return

    cabecalhos = list(contratos_faltando[0]["linha_completa"].keys())
    ws.append(cabecalhos)

    for item in contratos_faltando:
        linha = [item["linha_completa"].get(cab) for cab in cabecalhos]
        ws.append(linha)

    _aplicar_formatacao_tabela_pdf(ws)
    

def _atualizar_tabela_resumo_pdf(wb):
    """
    Atualiza a pequena tabela na aba Resumo, colunas G/H, com:
      G1:H1 → cabeçalho "Termo PDF"
      G2: "N op"     H2: contagem de Dados PDF (col A)
      G3: "V.CONF" H3: soma de Preço de Endosso da CCB (col E)
      G5: "Termo-Sistema"  H5: =E8 - H3
    """
    if "Resumo" not in wb.sheetnames or "Dados PDF" not in wb.sheetnames:
        return

    ws = wb["Resumo"]

    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(size=14, bold=True, color="000000")
    fonte_texto = Font(size=12, bold=True, color="000000")
    fill_titulo = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fmt_money = 'R$ #,##0.00'

    # Descobrir última linha de dados na aba Dados PDF (para fórmula de contagem/soma)
    ws_pdf = wb["Dados PDF"]
    ultima_linha_pdf = ws_pdf.max_row  # linha 1 = cabeçalho

    # Cabeçalho G1:H1
    ws.merge_cells("G1:H1")
    ws["G1"] = "Termo PDF"
    ws["G1"].font = fonte_titulo
    ws["G1"].fill = fill_titulo
    ws["G1"].alignment = alinhamento
    ws["G1"].border = borda
    ws["H1"].border = borda

    # G2:H2 — N op
    ws["G2"] = "N op"
    ws["H2"] = f"=COUNTA('Dados PDF'!A2:A{ultima_linha_pdf})"
    ws["G2"].font = fonte_texto
    ws["H2"].font = fonte_texto
    ws["G2"].alignment = alinhamento
    ws["H2"].alignment = alinhamento
    ws["G2"].border = borda
    ws["H2"].border = borda

    # G3:H3 — V.CONF (soma col E = Preço de Endosso)
    ws["G3"] = "V.CONF"
    ws["H3"] = f"=SUM('Dados PDF'!E2:E{ultima_linha_pdf})"
    ws["G3"].font = fonte_texto
    ws["H3"].font = fonte_texto
    ws["G3"].alignment = alinhamento
    ws["H3"].alignment = alinhamento
    ws["G3"].border = borda
    ws["H3"].border = borda
    ws["H3"].number_format = fmt_money

    # G5:H5 — Termo-Sistema = E8 (V.CONF do resumo operações) - H3 (V.CONF do PDF)
    ws["G5"] = "Termo-Sistema"
    ws["H5"] = "=E8-H3"
    ws["G5"].font = fonte_texto
    ws["H5"].font = fonte_texto
    ws["G5"].alignment = alinhamento
    ws["H5"].alignment = alinhamento
    ws["G5"].border = borda
    ws["H5"].border = borda
    ws["H5"].number_format = fmt_money

    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 22


def conferencia_pdf_parceiro_c(root, entry_data, progress_bar, event=None):
    """
    Terceiro botão PARCEIRO Z: Conferência Resumo (checagem) × PDF do Termo de Endosso.
    Pergunta se é o primeiro PDF para decidir se cria ou acumula dados.
    """
    # Pergunta "É o primeiro PDF?"
    resposta = messagebox.askyesno(
        "Conferência PDF",
        "É o primeiro PDF?\n\n"
        "• SIM → selecionar arquivo de checagem + PDF (início)\n"
        "• NÃO → selecionar arquivo de checagem existente + novo PDF (acumular)"
    )

    # Selecionar arquivo de checagem (resumo)
    caminho_resumo = filedialog.askopenfilename(
        title="Selecione o arquivo de Checagem PARCEIRO Z (.xlsx)",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_resumo:
        return

    # Selecionar PDF
    caminho_pdf = filedialog.askopenfilename(
        title="Selecione o PDF do Termo de Endosso",
        filetypes=[("PDF", "*.pdf")]
    )
    if not caminho_pdf:
        return

    try:
        progress_bar.pack(pady=10)
        progress_bar.set(0.10)
        root.update_idletasks()

        # Extrair dados do PDF
        registros_pdf = _extrair_dados_pdf_endosso(caminho_pdf)
        if not registros_pdf:
            raise Exception("Nenhum dado encontrado no PDF. Verifique se o arquivo é o Termo de Endosso correto.")

        progress_bar.set(0.35)
        root.update_idletasks()

        wb = load_workbook(caminho_resumo)

        # Criar/acumular aba "Dados PDF"
        if resposta:
            # Primeiro PDF: recriar aba
            if "Dados PDF" in wb.sheetnames:
                del wb["Dados PDF"]
            ws_pdf = wb.create_sheet("Dados PDF")
            _gravar_dados_pdf_na_aba(ws_pdf, registros_pdf, primeira_vez=True)
        else:
            # PDF adicional: acumular
            if "Dados PDF" not in wb.sheetnames:
                ws_pdf = wb.create_sheet("Dados PDF")
                _gravar_dados_pdf_na_aba(ws_pdf, registros_pdf, primeira_vez=True)
            else:
                ws_pdf = wb["Dados PDF"]
                _gravar_dados_pdf_na_aba(ws_pdf, registros_pdf, primeira_vez=False)

        _aplicar_formatacao_tabela_pdf(ws_pdf)

        progress_bar.set(0.55)
        root.update_idletasks()

        # Comparar com OP Tesouraria (se existir)
        contratos_faltando = []
        if "OP Tesouraria" in wb.sheetnames:
            contratos_faltando = _calcular_diferenca_ccbs(wb)
            _criar_aba_contratos_faltando(wb, contratos_faltando)

        progress_bar.set(0.75)
        root.update_idletasks()

        # Atualizar tabela na aba Resumo (colunas G/H)
        _atualizar_tabela_resumo_pdf(wb)

        mover_resumo_para_primeira_aba(wb)
        wb.save(caminho_resumo)

        progress_bar.set(1.0)
        root.update_idletasks()
        progress_bar.pack_forget()

        total_pdf = len(registros_pdf)
        msg = f"PDF processado com sucesso!\n\n{total_pdf} contrato(s) extraído(s) do PDF."
        if "OP Tesouraria" in wb.sheetnames:
            msg += f"\n\n{len(contratos_faltando)} contrato(s) da OP Tesouraria não encontrado(s) no PDF."
            if contratos_faltando:
               msg += "\nVeja a aba 'Contratos Faltantes PDF'."
        else:
            msg += "\n\n(Sem OP Tesouraria carregada — execute a Conferência Tesouraria para comparar.)"

        os.startfile(caminho_resumo)
        messagebox.showinfo("Sucesso", msg)

    except Exception as e:
        progress_bar.pack_forget()
        messagebox.showerror("Erro", str(e))
