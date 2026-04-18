# ============================================================================
# PROCESSAMENTO GERAL
# Funções base para resumo, dados originais e exportação de arquivos
# ============================================================================

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config.parceiros import COLUNAS_FINANCEIRAS, CODIGOS_GRUPO_D
from utils.excel import (
    aplicar_formatacao_padrao,
    obter_indice_coluna_por_nome,
    encontrar_indice_coluna_por_nomes,
)


def calcular_resumo(df, codigo_parceiro):
    dados_resumo = [
        ("Op.:", len(df)),
        ("V.B.:", df["VALOR_BRUTO"].sum()),
        ("V.L.:", df["VALOR_LIQUIDO"].sum()),
        ("IOF:", df["IOF"].sum())
    ]

    if codigo_parceiro in CODIGOS_GRUPO_D:
        df["AGIO"] = (df["VALOR_BRUTO"] * 0.0025).apply(lambda x: round(x + 1e-9, 2))
        df["VALOR_CESSAO"] = df["VALOR_LIQUIDO"] + df["IOF"] + df["AGIO"]
        dados_resumo.extend([
            ("ÁGIO:", df["AGIO"].sum()),
            ("V.CESSÃO:", df["VALOR_CESSAO"].sum())
        ])
    elif "CAD" in df.columns and df["CAD"].sum() > 0:
        dados_resumo.append(("CAD:", df["CAD"].sum()))

    return dados_resumo, df


def criar_aba_resumo(wb, dados_resumo):
    aba_resumo = wb.active
    aba_resumo.title = "Resumo"

    borda_fina = Side(style="thin", color="000000")
    borda_preta = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    fonte_titulo = Font(name="Aptos Narrow", size=16, bold=True, color="FFFFFF")
    fonte_valor = Font(name="Aptos Narrow", size=12, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")

    aba_resumo.merge_cells("B1:C1")
    aba_resumo["B1"] = "Resumo Sistema"
    aba_resumo["B1"].font = fonte_titulo
    aba_resumo["B1"].fill = PatternFill(start_color="215C98", end_color="215C98", fill_type="solid")
    aba_resumo["B1"].alignment = alinhamento
    aba_resumo["B1"].border = borda_preta
    aba_resumo["C1"].border = borda_preta

    for i, (rotulo, valor) in enumerate(dados_resumo, start=2):
        cel_rotulo = aba_resumo.cell(row=i, column=2, value=rotulo)
        cel_rotulo.font = fonte_valor
        cel_rotulo.alignment = alinhamento
        cel_rotulo.border = borda_preta

        cel_valor = aba_resumo.cell(row=i, column=3, value=valor)
        cel_valor.font = fonte_valor
        cel_valor.alignment = alinhamento
        cel_valor.border = borda_preta

        if rotulo != "Op.:":
            cel_valor.number_format = 'R$ #,##0.00'

    aba_resumo.column_dimensions["B"].width = 18
    aba_resumo.column_dimensions["C"].width = 22


def criar_aba_dados_originais(wb, df):
    if "Dados Originais" in wb.sheetnames:
        del wb["Dados Originais"]

    aba_dados = wb.create_sheet(title="Dados Originais")

    for linha in dataframe_to_rows(df, index=False, header=True):
        aba_dados.append(linha)

    aplicar_formatacao_padrao(aba_dados, COLUNAS_FINANCEIRAS)
    
def gerar_arquivo_exportacao(df, pasta_saida, nome_parceiro, codigo_parceiro, data_digitada):
    codigo_seguro = codigo_parceiro.replace("/", "_")
    if nome_parceiro == "PARCEIRO A":
        nome_arquivo = f"PARCEIRO A - {data_digitada}.xlsx"
    else:
        nome_arquivo = f"{nome_parceiro} {codigo_seguro} - {data_digitada}.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    wb_exportacao = Workbook()
    aba_exportacao = wb_exportacao.active

    for linha in dataframe_to_rows(df, index=False, header=True):
        aba_exportacao.append(linha)

    colunas_para_somar = ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD"]
    mapa_colunas = {}

    for idx, cell in enumerate(aba_exportacao[1], start=1):
        mapa_colunas[cell.value] = idx

    linha_total = aba_exportacao.max_row + 1
    aba_exportacao.cell(row=linha_total, column=1, value="TOTAL")

    for nome_coluna in colunas_para_somar:
        if nome_coluna in mapa_colunas:
            indice_coluna = mapa_colunas[nome_coluna]
            soma = df[nome_coluna].sum()
            aba_exportacao.cell(row=linha_total, column=indice_coluna, value=soma)

    aplicar_formatacao_padrao(aba_exportacao, COLUNAS_FINANCEIRAS)

    fill_total = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fonte_total = Font(bold=True)
    borda_fina = Side(style="thin", color="000000")
    borda_total = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")

    for col in range(1, aba_exportacao.max_column + 1):
        cell = aba_exportacao.cell(row=linha_total, column=col)
        cell.fill = fill_total
        cell.font = fonte_total
        cell.border = borda_total
        cell.alignment = alinhamento

        if aba_exportacao.cell(row=1, column=col).value in COLUNAS_FINANCEIRAS:
            cell.number_format = 'R$ #,##0.00'

    wb_exportacao.save(caminho_saida)    