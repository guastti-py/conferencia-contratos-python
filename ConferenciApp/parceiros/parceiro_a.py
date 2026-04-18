from tkinter import messagebox, filedialog
# ==================================================================================
# PARCEIRO: PARCEIRO A
# Regras de geração de resumo e processamento dos arquivos PF e PJ
# ==================================================================================

import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from utils.arquivos import (
    validar_data,
    carregar_arquivo,
    padronizar_colunas,
    converter_colunas_numericas,
)
from utils.excel import (
    obter_indice_coluna_por_nome,
    encontrar_indice_coluna_por_nomes,
)
from parceiros.geral import (
    criar_aba_dados_originais,
    gerar_arquivo_exportacao,
)

def criar_aba_resumo_parceiro_a(wb):
    if "Resumo" in wb.sheetnames:
        aba_resumo = wb["Resumo"]
        if aba_resumo != wb.active:
            del wb["Resumo"]
            aba_resumo = wb.active
            aba_resumo.title = "Resumo"
        else:
            aba_resumo.delete_rows(1, aba_resumo.max_row)
            aba_resumo.title = "Resumo"
    else:
        aba_resumo = wb.active
        aba_resumo.title = "Resumo"

    ws_dados = wb["Dados Originais"]

    col_contrato = encontrar_indice_coluna_por_nomes(ws_dados, ["CONTRATO", "Nº DA CCB", "NUM_CONTRATO", "CCB"])
    col_valor_bruto = obter_indice_coluna_por_nome(ws_dados, "VALOR_BRUTO")
    col_valor_liquido = obter_indice_coluna_por_nome(ws_dados, "VALOR_LIQUIDO")
    col_iof = obter_indice_coluna_por_nome(ws_dados, "IOF")
    col_cad = obter_indice_coluna_por_nome(ws_dados, "CAD")
    col_cpf = obter_indice_coluna_por_nome(ws_dados, "CPF")
    col_cnpj = obter_indice_coluna_por_nome(ws_dados, "CNPJ")

    if col_contrato is None:
        col_contrato = 1

    if col_valor_bruto is None or col_valor_liquido is None or col_iof is None:
        raise Exception("Não encontrei as colunas financeiras obrigatórias em 'Dados Originais'.")

    letra_contrato = get_column_letter(col_contrato)
    letra_vb = get_column_letter(col_valor_bruto)
    letra_vl = get_column_letter(col_valor_liquido)
    letra_iof = get_column_letter(col_iof)
    letra_cad = get_column_letter(col_cad) if col_cad else None
    letra_cpf = get_column_letter(col_cpf) if col_cpf else None
    letra_cnpj = get_column_letter(col_cnpj) if col_cnpj else None

    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")

    fonte_titulo = Font(name="Aptos Narrow", size=16, bold=True, color="000000")
    fonte_subtitulo = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    fonte_valor = Font(name="Aptos Narrow", size=12)
    alinhamento = Alignment(horizontal="center", vertical="center")

    cor_sistema = "FFC000"
    cor_total = "92D050"
    cor_pf = "FFFF00"
    cor_pj = "83CCEB"

    aba_resumo.column_dimensions["B"].width = 12
    aba_resumo.column_dimensions["C"].width = 22

    aba_resumo.merge_cells("B1:C1")
    cel_sistema = aba_resumo["B1"]
    cel_sistema.value = "Sistema"
    cel_sistema.font = fonte_titulo
    cel_sistema.fill = PatternFill(start_color=cor_sistema, end_color=cor_sistema, fill_type="solid")
    cel_sistema.alignment = alinhamento

    aba_resumo["B1"].border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa
    )
    aba_resumo["C1"].border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa
    )

    def escrever_titulo_bloco(linha_inicial, titulo_bloco, cor_cabecalho):
        aba_resumo.merge_cells(start_row=linha_inicial, start_column=2, end_row=linha_inicial, end_column=3)

        cel_titulo = aba_resumo.cell(row=linha_inicial, column=2, value=titulo_bloco)
        cel_titulo.font = fonte_subtitulo
        cel_titulo.fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
        cel_titulo.alignment = alinhamento

        aba_resumo.cell(row=linha_inicial, column=2).border = Border(
            left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_fina
        )
        aba_resumo.cell(row=linha_inicial, column=3).border = Border(
            left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_fina
        )

    def escrever_linha_formula(linha, rotulo, formula, ultima=False):
        cel_rotulo = aba_resumo.cell(row=linha, column=2, value=rotulo)
        cel_rotulo.font = fonte_valor
        cel_rotulo.alignment = alinhamento

        cel_valor = aba_resumo.cell(row=linha, column=3, value=formula)
        cel_valor.font = fonte_valor
        cel_valor.alignment = alinhamento

        if rotulo != "Op.:":
            cel_valor.number_format = 'R$ #,##0.00'

        aba_resumo.cell(row=linha, column=2).border = Border(
            left=borda_grossa,
            right=borda_fina,
            top=borda_fina,
            bottom=borda_grossa if ultima else borda_fina
        )
        aba_resumo.cell(row=linha, column=3).border = Border(
            left=borda_fina,
            right=borda_grossa,
            top=borda_fina,
            bottom=borda_grossa if ultima else borda_fina
        )

    escrever_titulo_bloco(3, "TOTAL", cor_total)
    escrever_linha_formula(4, "Op.:", f'=COUNTA(\'Dados Originais\'!{letra_contrato}:{letra_contrato})-1')
    escrever_linha_formula(5, "V.B.:", f'=SUM(\'Dados Originais\'!{letra_vb}:{letra_vb})')
    escrever_linha_formula(6, "V.L.:", f'=SUM(\'Dados Originais\'!{letra_vl}:{letra_vl})')
    escrever_linha_formula(7, "IOF:", f'=SUM(\'Dados Originais\'!{letra_iof}:{letra_iof})')
    if letra_cad:
        escrever_linha_formula(8, "CAD:", f'=SUM(\'Dados Originais\'!{letra_cad}:{letra_cad})', ultima=True)
    else:
        aba_resumo.cell(row=8, column=2).border = Border(left=borda_grossa, right=borda_fina, top=borda_fina, bottom=borda_grossa)
        aba_resumo.cell(row=8, column=3).border = Border(left=borda_fina, right=borda_grossa, top=borda_fina, bottom=borda_grossa)

    escrever_titulo_bloco(10, "Pag Seguro PF", cor_pf)
    if letra_cpf:
        escrever_linha_formula(11, "Op.:", f'=COUNTIFS(\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>",\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>CPF")')
        escrever_linha_formula(12, "V.B.:", f'=SUMIFS(\'Dados Originais\'!{letra_vb}:{letra_vb},\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>",\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>CPF")')
        escrever_linha_formula(13, "V.L.:", f'=SUMIFS(\'Dados Originais\'!{letra_vl}:{letra_vl},\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>",\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>CPF")')
        escrever_linha_formula(14, "IOF:", f'=SUMIFS(\'Dados Originais\'!{letra_iof}:{letra_iof},\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>",\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>CPF")')
        if letra_cad:
            escrever_linha_formula(15, "CAD:", f'=SUMIFS(\'Dados Originais\'!{letra_cad}:{letra_cad},\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>",\'Dados Originais\'!{letra_cpf}:{letra_cpf},"<>CPF")', ultima=True)
        else:
            aba_resumo.cell(row=15, column=2).border = Border(left=borda_grossa, right=borda_fina, top=borda_fina, bottom=borda_grossa)
            aba_resumo.cell(row=15, column=3).border = Border(left=borda_fina, right=borda_grossa, top=borda_fina, bottom=borda_grossa)

    escrever_titulo_bloco(17, "Pag Seguro PJ", cor_pj)
    if letra_cnpj:
        escrever_linha_formula(18, "Op.:", f'=COUNTIFS(\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>",\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>CNPJ")')
        escrever_linha_formula(19, "V.B.:", f'=SUMIFS(\'Dados Originais\'!{letra_vb}:{letra_vb},\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>",\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>CNPJ")')
        escrever_linha_formula(20, "V.L.:", f'=SUMIFS(\'Dados Originais\'!{letra_vl}:{letra_vl},\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>",\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>CNPJ")')
        escrever_linha_formula(21, "IOF:", f'=SUMIFS(\'Dados Originais\'!{letra_iof}:{letra_iof},\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>",\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>CNPJ")')
        if letra_cad:
            escrever_linha_formula(22, "CAD:", f'=SUMIFS(\'Dados Originais\'!{letra_cad}:{letra_cad},\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>",\'Dados Originais\'!{letra_cnpj}:{letra_cnpj},"<>CNPJ")', ultima=True)
        else:
            aba_resumo.cell(row=22, column=2).border = Border(left=borda_grossa, right=borda_fina, top=borda_fina, bottom=borda_grossa)
            aba_resumo.cell(row=22, column=3).border = Border(left=borda_fina, right=borda_grossa, top=borda_fina, bottom=borda_grossa)


def selecionar_e_processar_parceiro_a(root, entry_data, progress_bar, event=None):
    data_digitada = entry_data.get().strip()

    if not validar_data(data_digitada):
        messagebox.showwarning("Atenção", "Por favor, digite uma data válida no formato dd.mm.aaaa.")
        return

    caminho_arquivo_pf = filedialog.askopenfilename(
        title="Selecionar arquivo PARCEIRO A PF",
        filetypes=[("Excel ou CSV", "*.xlsx *.xls *.csv")]
    )
    if not caminho_arquivo_pf:
        return

    caminho_arquivo_pj = filedialog.askopenfilename(
        title="Selecionar arquivo PARCEIRO A PJ",
        filetypes=[("Excel ou CSV", "*.xlsx *.xls *.csv")]
    )
    if not caminho_arquivo_pj:
        return

    try:
        progress_bar.pack(pady=10)
        progress_bar.set(0.20)
        root.update_idletasks()

        df_pf = carregar_arquivo(caminho_arquivo_pf)
        df_pj = carregar_arquivo(caminho_arquivo_pj)

        df_pf = padronizar_colunas(df_pf)
        df_pj = padronizar_colunas(df_pj)

        df_pf = converter_colunas_numericas(df_pf)
        df_pj = converter_colunas_numericas(df_pj)

        df_unificado = pd.concat([df_pf, df_pj], ignore_index=True)

        progress_bar.set(0.45)
        root.update_idletasks()

        pasta_saida = os.path.dirname(caminho_arquivo_pf)

        data_formatada = data_digitada.replace("/", ".")
        nome_arquivo_resumo = f"Checagem PARCEIRO A PF e PJ - {data_formatada}.xlsx"

        caminho_resumo = os.path.join(pasta_saida, nome_arquivo_resumo)

        wb_resumo = Workbook()

        ws_inicial = wb_resumo.active
        ws_inicial.title = "Resumo"

        criar_aba_dados_originais(wb_resumo, df_unificado)
        criar_aba_resumo_parceiro_a(wb_resumo)

        wb_resumo.save(caminho_resumo)

        progress_bar.set(0.80)
        root.update_idletasks()

        gerar_arquivo_exportacao(
            df_unificado,
            pasta_saida,
            "PARCEIRO A PF e PJ",
            "114/115",
            data_digitada
        )

        progress_bar.set(1.0)
        root.update_idletasks()
        progress_bar.pack_forget()

        os.startfile(pasta_saida)
        messagebox.showinfo("Sucesso", "Parceiro A processado com os dois arquivos unidos!")

    except Exception as erro:
        progress_bar.pack_forget()
        messagebox.showerror("Erro", str(erro))

def carregar_csv_conferencia_parceiro_a(caminho_csv):
    df = pd.read_csv(
        caminho_csv,
        sep=";",
        encoding="utf-8-sig",
        header=2,
        usecols=list(range(8))
    )

    df.columns = [str(col).strip() for col in df.columns]

    mapa_renomear = {
        "Número da CCB": "CONTRATO",
        "Id": "ID",
        "Data de contrato": "DATA_CONTRATO",
        "Cnpj": "CNPJ",
        "Valor bruto": "VALOR_BRUTO",
        "Iof": "IOF",
        "Comissão": "CAD",
        "Valor líquido": "VALOR_LIQUIDO"
    }

    df = df.rename(columns=mapa_renomear)

    if "CNPJ" in df.columns:
        df["CNPJ"] = df["CNPJ"].fillna("").astype(str).str.strip()
        df.loc[df["CNPJ"].isin(["nan", "None", "NaN", ""]), "CNPJ"] = "1"
        df.loc[~df["CNPJ"].isin(["1"]), "CNPJ"] = "2"

    for col in ["VALOR_BRUTO", "IOF", "CAD", "VALOR_LIQUIDO"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("R$", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
                .str.strip()
            )
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def salvar_aba_conferencia_parceiro_a(wb, df_conf):
    from utils.excel import aplicar_formatacao_padrao
    nome_aba = "Dados Conferência"

    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws_conf = wb.create_sheet(title=nome_aba)

    from openpyxl.utils.dataframe import dataframe_to_rows
    for linha in dataframe_to_rows(df_conf, index=False, header=True):
        ws_conf.append(linha)

    aplicar_formatacao_padrao(ws_conf, ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD"])


def escrever_bloco_conferencia_parceiro_a(ws, linha_inicial, coluna_inicial, titulo_bloco, formulas_bloco, cor_cabecalho):
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")

    fonte_subtitulo = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    fonte_valor = Font(name="Aptos Narrow", size=12)
    alinhamento = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=linha_inicial,
        start_column=coluna_inicial,
        end_row=linha_inicial,
        end_column=coluna_inicial + 1
    )

    cel_titulo = ws.cell(row=linha_inicial, column=coluna_inicial, value=titulo_bloco)
    cel_titulo.font = fonte_subtitulo
    cel_titulo.fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
    cel_titulo.alignment = alinhamento

    ws.cell(row=linha_inicial, column=coluna_inicial).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_fina
    )
    ws.cell(row=linha_inicial, column=coluna_inicial + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_fina
    )

    total_linhas = len(formulas_bloco)

    for i, (rotulo, formula) in enumerate(formulas_bloco, start=1):
        linha = linha_inicial + i
        eh_ultima = i == total_linhas

        cel_rotulo = ws.cell(row=linha, column=coluna_inicial, value=rotulo)
        cel_rotulo.font = fonte_valor
        cel_rotulo.alignment = alinhamento

        cel_valor = ws.cell(row=linha, column=coluna_inicial + 1, value=formula)
        cel_valor.font = fonte_valor
        cel_valor.alignment = alinhamento

        if rotulo != "Op.:":
            cel_valor.number_format = 'R$ #,##0.00'

        ws.cell(row=linha, column=coluna_inicial).border = Border(
            left=borda_grossa,
            right=borda_fina,
            top=borda_fina,
            bottom=borda_grossa if eh_ultima else borda_fina
        )
        ws.cell(row=linha, column=coluna_inicial + 1).border = Border(
            left=borda_fina,
            right=borda_grossa,
            top=borda_fina,
            bottom=borda_grossa if eh_ultima else borda_fina
        )


def escrever_bloco_checagem(ws, linha_inicial, coluna_inicial, titulo, linha_base_sistema, linha_base_conferencia, cor_titulo):
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")

    fonte = Font(name="Aptos Narrow", size=11)
    fonte_titulo = Font(name="Aptos Narrow", size=12, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")

    itens = ["Op.:", "V.B.:", "V.L.:", "IOF:", "CAD:"]

    ws.merge_cells(
        start_row=linha_inicial,
        start_column=coluna_inicial,
        end_row=linha_inicial,
        end_column=coluna_inicial + 1
    )

    cel_titulo = ws.cell(row=linha_inicial, column=coluna_inicial, value=titulo)
    cel_titulo.font = fonte_titulo
    cel_titulo.fill = PatternFill(start_color=cor_titulo, end_color=cor_titulo, fill_type="solid")
    cel_titulo.alignment = alinhamento

    ws.cell(row=linha_inicial, column=coluna_inicial).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_fina
    )
    ws.cell(row=linha_inicial, column=coluna_inicial + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_fina
    )

    for i, rotulo in enumerate(itens, start=1):
        linha = linha_inicial + i
        eh_ultima = i == len(itens)

        cel_rotulo = ws.cell(row=linha, column=coluna_inicial, value=rotulo)
        cel_rotulo.font = fonte
        cel_rotulo.alignment = alinhamento

        formula = f'=IF(C{linha_base_sistema + i}=G{linha_base_conferencia + i},"CONFERE","NÃO CONFERE")'
        cel_resultado = ws.cell(row=linha, column=coluna_inicial + 1, value=formula)
        cel_resultado.font = fonte
        cel_resultado.alignment = alinhamento

        ws.cell(row=linha, column=coluna_inicial).border = Border(
            left=borda_grossa,
            right=borda_fina,
            top=borda_fina,
            bottom=borda_grossa if eh_ultima else borda_fina
        )
        ws.cell(row=linha, column=coluna_inicial + 1).border = Border(
            left=borda_fina,
            right=borda_grossa,
            top=borda_fina,
            bottom=borda_grossa if eh_ultima else borda_fina
        )


def ajustar_dados_originais_com_conferencia(wb, caminho_excel, df_conf):
    from utils.excel import aplicar_formatacao_padrao
    from config.parceiros import COLUNAS_FINANCEIRAS
    from openpyxl.utils.dataframe import dataframe_to_rows

    if "Dados Originais" not in wb.sheetnames:
        raise Exception("A aba 'Dados Originais' não foi encontrada no Excel.")

    df_orig = pd.read_excel(caminho_excel, sheet_name="Dados Originais")
    df_orig.columns = df_orig.columns.str.strip()

    if "CONTRATO" not in df_conf.columns:
        raise Exception("A coluna 'CONTRATO' não foi encontrada nos Dados Conferência.")

    col_contrato_orig = None
    for col in df_orig.columns:
        nome = str(col).upper().strip()
        if "CONTRATO" in nome or "CCB" in nome:
            col_contrato_orig = col
            break

    if col_contrato_orig is None:
        raise Exception("Não encontrei a coluna de contrato na aba 'Dados Originais'.")

    contratos_conf = set(df_conf["CONTRATO"].astype(str).str.strip())
    contratos_orig = df_orig[col_contrato_orig].astype(str).str.strip()

    df_faltantes = df_orig[~contratos_orig.isin(contratos_conf)].copy()
    df_orig_filtrado = df_orig[contratos_orig.isin(contratos_conf)].copy()

    if "Dados Originais" in wb.sheetnames:
        del wb["Dados Originais"]
    ws_nova_orig = wb.create_sheet("Dados Originais")

    for linha in dataframe_to_rows(df_orig_filtrado, index=False, header=True):
        ws_nova_orig.append(linha)

    aplicar_formatacao_padrao(ws_nova_orig, COLUNAS_FINANCEIRAS)

    nome_aba_faltantes = "Contratos Não Conf."
    if nome_aba_faltantes in wb.sheetnames:
        del wb[nome_aba_faltantes]
    ws_faltantes = wb.create_sheet(nome_aba_faltantes)

    for linha in dataframe_to_rows(df_faltantes, index=False, header=True):
        ws_faltantes.append(linha)

    aplicar_formatacao_padrao(ws_faltantes, COLUNAS_FINANCEIRAS)

    contratos_validos = set(df_orig_filtrado[col_contrato_orig].astype(str).str.strip())
    df_conf_filtrado = df_conf[
        df_conf["CONTRATO"].astype(str).str.strip().isin(contratos_validos)
    ].copy()

    return df_conf_filtrado, df_orig_filtrado


def atualizar_segundo_arquivo_filtrado(caminho_excel_resumo, df_orig_filtrado):
    from utils.excel import aplicar_formatacao_padrao
    from config.parceiros import COLUNAS_FINANCEIRAS
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook

    pasta_saida = os.path.dirname(caminho_excel_resumo)

    nome_arquivo_exportacao = "PARCEIRO A PF e PJ - " + os.path.basename(caminho_excel_resumo).replace(
        "Checagem PARCEIRO A PF e PJ - ", ""
    )

    caminho_exportacao = os.path.join(pasta_saida, nome_arquivo_exportacao)

    wb_exportacao = Workbook()
    aba_exportacao = wb_exportacao.active

    for linha in dataframe_to_rows(df_orig_filtrado, index=False, header=True):
        aba_exportacao.append(linha)

    colunas_para_somar = ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD"]
    mapa_colunas = {}

    for idx, cell in enumerate(aba_exportacao[1], start=1):
        mapa_colunas[cell.value] = idx

    linha_total = aba_exportacao.max_row + 1
    aba_exportacao.cell(row=linha_total, column=1, value="TOTAL")

    for nome_coluna in colunas_para_somar:
        if nome_coluna in mapa_colunas and nome_coluna in df_orig_filtrado.columns:
            indice_coluna = mapa_colunas[nome_coluna]
            soma = df_orig_filtrado[nome_coluna].sum()
            aba_exportacao.cell(row=linha_total, column=indice_coluna, value=soma)

    aplicar_formatacao_padrao(aba_exportacao, COLUNAS_FINANCEIRAS)

    fill_total = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fonte_total = Font(bold=True)
    borda_fina = Side(style="thin", color="000000")
    borda_total = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")

    for col in range(1, aba_exportacao.max_column + 1):
        cell = aba_exportacao.cell(row=linha_total, column=col)
        cell.fill = fill_total
        cell.font = fonte_total
        cell.border = borda_total
        cell.alignment = alinhamento

        if aba_exportacao.cell(row=1, column=col).value in COLUNAS_FINANCEIRAS:
            cell.number_format = 'R$ #,##0.00'

    wb_exportacao.save(caminho_exportacao)


def fazer_conferencia_parceiro_a():
    from openpyxl import load_workbook
    from utils.excel import aplicar_formatacao_condicional_checagem

    messagebox.showinfo(
        "Conferência Parceiro A",
        "1º Selecione o arquivo EXCEL de Resumo\n2º Selecione o CSV de conferência do Parceiro A"
    )

    caminho_excel = filedialog.askopenfilename(
        title="Selecionar Resumo Excel",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_excel:
        return

    caminho_csv = filedialog.askopenfilename(
        title="Selecionar CSV de Conferência Parceiro A",
        filetypes=[("CSV", "*.csv")]
    )
    if not caminho_csv:
        return

    try:
        wb = load_workbook(caminho_excel)

        if "Resumo" not in wb.sheetnames:
            raise Exception("A aba 'Resumo' não foi encontrada no Excel.")

        ws_resumo = wb["Resumo"]

        df_conf = carregar_csv_conferencia_parceiro_a(caminho_csv)

        df_conf_filtrado, df_orig_filtrado = ajustar_dados_originais_com_conferencia(wb, caminho_excel, df_conf)
        salvar_aba_conferencia_parceiro_a(wb, df_conf_filtrado)

        ws_conf = wb["Dados Conferência"]

        col_contrato = encontrar_indice_coluna_por_nomes(ws_conf, ["CONTRATO", "Nº DA CCB", "NUM_CONTRATO", "CCB"])
        col_vb = obter_indice_coluna_por_nome(ws_conf, "VALOR_BRUTO")
        col_vl = obter_indice_coluna_por_nome(ws_conf, "VALOR_LIQUIDO")
        col_iof = obter_indice_coluna_por_nome(ws_conf, "IOF")
        col_cad = obter_indice_coluna_por_nome(ws_conf, "CAD")
        col_cnpj = obter_indice_coluna_por_nome(ws_conf, "CNPJ")

        if col_contrato is None or col_vb is None or col_vl is None or col_iof is None or col_cnpj is None:
            raise Exception("Não encontrei todas as colunas necessárias na aba 'Dados Conferência'.")

        letra_contrato = get_column_letter(col_contrato)
        letra_vb = get_column_letter(col_vb)
        letra_vl = get_column_letter(col_vl)
        letra_iof = get_column_letter(col_iof)
        letra_cad = get_column_letter(col_cad) if col_cad else None
        letra_cnpj = get_column_letter(col_cnpj)

        formulas_total = [
            ("Op.:", f'=COUNTA(\'Dados Conferência\'!{letra_contrato}:{letra_contrato})-1'),
            ("V.B.:", f'=SUM(\'Dados Conferência\'!{letra_vb}:{letra_vb})'),
            ("V.L.:", f'=SUM(\'Dados Conferência\'!{letra_vl}:{letra_vl})'),
            ("IOF:", f'=SUM(\'Dados Conferência\'!{letra_iof}:{letra_iof})'),
            ("CAD:", f'=SUM(\'Dados Conferência\'!{letra_cad}:{letra_cad})' if letra_cad else '=0')
        ]

        formulas_pf = [
            ("Op.:", f'=COUNTIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"1")'),
            ("V.B.:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"1",\'Dados Conferência\'!{letra_vb}:{letra_vb})'),
            ("V.L.:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"1",\'Dados Conferência\'!{letra_vl}:{letra_vl})'),
            ("IOF:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"1",\'Dados Conferência\'!{letra_iof}:{letra_iof})'),
            ("CAD:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"1",\'Dados Conferência\'!{letra_cad}:{letra_cad})' if letra_cad else '=0')
        ]

        formulas_pj = [
            ("Op.:", f'=COUNTIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"2")'),
            ("V.B.:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"2",\'Dados Conferência\'!{letra_vb}:{letra_vb})'),
            ("V.L.:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"2",\'Dados Conferência\'!{letra_vl}:{letra_vl})'),
            ("IOF:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"2",\'Dados Conferência\'!{letra_iof}:{letra_iof})'),
            ("CAD:", f'=SUMIF(\'Dados Conferência\'!{letra_cnpj}:{letra_cnpj},"2",\'Dados Conferência\'!{letra_cad}:{letra_cad})' if letra_cad else '=0')
        ]

        ws_resumo.column_dimensions["F"].width = 12
        ws_resumo.column_dimensions["G"].width = 22
        ws_resumo.column_dimensions["J"].width = 12
        ws_resumo.column_dimensions["K"].width = 18

        borda_fina = Side(style="thin", color="000000")
        borda_grossa = Side(style="medium", color="000000")

        ws_resumo.merge_cells("F1:G1")
        cel_conf = ws_resumo["F1"]
        cel_conf.value = "PAG SEGURO"
        cel_conf.font = Font(name="Aptos Narrow", size=16, bold=True, color="000000")
        cel_conf.fill = PatternFill(start_color="FCE76A", end_color="FCE76A", fill_type="solid")
        cel_conf.alignment = Alignment(horizontal="center", vertical="center")

        ws_resumo["F1"].border = Border(
            left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa
        )
        ws_resumo["G1"].border = Border(
            left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa
        )

        escrever_bloco_conferencia_parceiro_a(ws_resumo, 3, 6, "TOTAL", formulas_total, "92D050")
        escrever_bloco_conferencia_parceiro_a(ws_resumo, 10, 6, "Pag Seguro PF", formulas_pf, "FFFF00")
        escrever_bloco_conferencia_parceiro_a(ws_resumo, 17, 6, "Pag Seguro PJ", formulas_pj, "83CCEB")

        ws_resumo.merge_cells("J1:K1")
        ws_resumo["J1"] = "Checagem"
        ws_resumo["J1"].font = Font(name="Aptos Narrow", size=16, bold=True, color="000000")
        ws_resumo["J1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo["J1"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws_resumo["J1"].border = Border(
            left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa
        )
        ws_resumo["K1"].border = Border(
            left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa
        )

        escrever_bloco_checagem(ws_resumo, 3, 10, "TOTAL", 3, 3, "92D050")
        escrever_bloco_checagem(ws_resumo, 10, 10, "Pag Seguro PF", 10, 10, "FFFF00")
        escrever_bloco_checagem(ws_resumo, 17, 10, "Pag Seguro PJ", 17, 17, "83CCEB")

        aplicar_formatacao_condicional_checagem(ws_resumo)

        atualizar_segundo_arquivo_filtrado(caminho_excel, df_orig_filtrado)

        try:
            wb.save(caminho_excel)
        except PermissionError:
            messagebox.showerror(
                "Arquivo aberto",
                f"Feche o arquivo no Excel e tente novamente:\n{caminho_excel}"
            )
            return
        messagebox.showinfo("Sucesso", "Conferência do Parceiro A salva no Excel!")
        os.startfile(caminho_excel)

    except Exception as erro:
        messagebox.showerror("Erro na Conferência Parceiro A", str(erro))
