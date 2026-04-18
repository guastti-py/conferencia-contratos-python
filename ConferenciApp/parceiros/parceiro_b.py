from tkinter import messagebox, filedialog
# ==============================================================================
# PARCEIRO: PARCEIRO B
# Regras de resumo, conferencia CSV e confêrencia de endosso em PDF
# ==============================================================================

import os
import re
import pandas as pd
import pdfplumber

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from utils.arquivos import (
    converter_colunas_numericas,
    validar_data,
    carregar_arquivo,
    padronizar_colunas,
    limpar_valor_monetario,
)
from utils.excel import (
    aplicar_formatacao_padrao,
    obter_indice_coluna_por_nome,
    encontrar_indice_coluna_por_nomes,
)
from parceiros.geral import criar_aba_dados_originais


PARCEIRO_B_NOMES_CURTOS = {
    "129": "GRUPO B2",
    "270": "GRUPO B1",
    "281": "GRUPO B3",
    "264": "FIDC I",
    "265": "FIDC II",
    "297": "FIDC III",
}


def obter_nome_curto_parceiro_b(codigo_parceiro, nome_padrao=""):
    return PARCEIRO_B_NOMES_CURTOS.get(codigo_parceiro, nome_padrao.strip() or codigo_parceiro)


def obter_nome_resumo_parceiro_b(codigo_parceiro, data_digitada, nome_padrao=""):
    nome_curto = obter_nome_curto_parceiro_b(codigo_parceiro, nome_padrao)
    return f"Checagem - {nome_curto} - {data_digitada}.xlsx"


def obter_nome_exportacao_parceiro_b(codigo_parceiro, data_digitada, nome_padrao=""):
    nome_curto = obter_nome_curto_parceiro_b(codigo_parceiro, nome_padrao)
    return f"{nome_curto} - {data_digitada}.xlsx"


def salvar_planilha_parceiro_b(df, caminho_saida):
    wb_exportacao = Workbook()
    aba_exportacao = wb_exportacao.active

    for linha in dataframe_to_rows(df, index=False, header=True):
        aba_exportacao.append(linha)

    colunas_para_somar = ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD"]
    mapa_colunas = {}

    for idx, cell in enumerate(aba_exportacao[1], start=1):
        mapa_colunas[cell.value] = idx

    linha_total = aba_exportacao.max_row + 1
    aba_exportacao.cell(row=linha_total, column=1, value="TOTAL")

    for nome_coluna in colunas_para_somar:
        if nome_coluna in mapa_colunas:
            indice_coluna = mapa_colunas[nome_coluna]
            soma = df[nome_coluna].sum()
            aba_exportacao.cell(row=linha_total, column=indice_coluna, value=soma)

    aplicar_formatacao_padrao(aba_exportacao, ["VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD", "AGIO", "VALOR_CESSAO"])

    fill_total = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fonte_total = Font(bold=True)
    borda_fina = Side(style="thin", color="000000")
    borda_total = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")

    for col in range(1, aba_exportacao.max_column + 1):
        cell = aba_exportacao.cell(row=linha_total, column=col)
        cell.fill = fill_total
        cell.font = fonte_total
        cell.border = borda_total
        cell.alignment = alinhamento

        if aba_exportacao.cell(row=1, column=col).value in {"VALOR_BRUTO", "VALOR_LIQUIDO", "IOF", "CAD", "AGIO", "VALOR_CESSAO"}:
            cell.number_format = "R$ #,##0.00"

    wb_exportacao.save(caminho_saida)


def criar_aba_dados_parceiro_parceiro_b(wb, df_parceiro):
    if "Dados Parceiro" in wb.sheetnames:
        del wb["Dados Parceiro"]

    ws = wb.create_sheet(title="Dados Parceiro")

    for linha in dataframe_to_rows(df_parceiro, index=False, header=True):
        ws.append(linha)

    aplicar_formatacao_padrao(ws, [])


def _estilizar_bloco_resumo(ws, linha_inicio, titulo, cor_titulo, linhas, altura_titulo=22):
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(name="Aptos Narrow", size=14, bold=True, color="000000")
    fonte_texto = Font(name="Aptos Narrow", size=12, bold=False, color="000000")

    ws.merge_cells(start_row=linha_inicio, start_column=2, end_row=linha_inicio, end_column=3)
    cel_titulo = ws.cell(row=linha_inicio, column=2, value=titulo)
    cel_titulo.font = fonte_titulo
    cel_titulo.fill = PatternFill(start_color=cor_titulo, end_color=cor_titulo, fill_type="solid")
    cel_titulo.alignment = alinhamento
    ws.row_dimensions[linha_inicio].height = altura_titulo

    for col in (2, 3):
        ws.cell(row=linha_inicio, column=col).border = borda

    for idx, (rotulo, valor) in enumerate(linhas, start=1):
        linha = linha_inicio + idx
        c1 = ws.cell(row=linha, column=2, value=rotulo)
        c2 = ws.cell(row=linha, column=3, value=valor)
        c1.font = fonte_texto
        c2.font = fonte_texto
        c1.alignment = alinhamento
        c2.alignment = alinhamento
        c1.border = borda
        c2.border = borda
        if rotulo != "Op.:":
            c2.number_format = 'R$ #,##0.00'

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22


def criar_aba_resumo_parceiro_b_sistema(wb):
    ws = wb.active
    ws.title = "Resumo"
    if ws.max_row > 1 or ws.max_column > 1 or ws["A1"].value is not None:
        ws.delete_rows(1, ws.max_row)

    ws_dados = wb["Dados Originais"]

    col_contrato = encontrar_indice_coluna_por_nomes(ws_dados, ["CONTRATO", "Nº DA CCB", "NUM_CONTRATO", "CCB"]) or 1
    col_vb = obter_indice_coluna_por_nome(ws_dados, "VALOR_BRUTO")
    col_vl = obter_indice_coluna_por_nome(ws_dados, "VALOR_LIQUIDO")
    col_iof = obter_indice_coluna_por_nome(ws_dados, "IOF")
    col_cad = obter_indice_coluna_por_nome(ws_dados, "CAD")

    if not all([col_vb, col_vl, col_iof, col_cad]):
        raise Exception("Não encontrei as colunas VALOR_BRUTO, VALOR_LIQUIDO, IOF e CAD em 'Dados Originais'.")

    letra_contrato = get_column_letter(col_contrato)
    letra_vb = get_column_letter(col_vb)
    letra_vl = get_column_letter(col_vl)
    letra_iof = get_column_letter(col_iof)
    letra_cad = get_column_letter(col_cad)

    linhas = [
        ("Op.:", f"=COUNTA('Dados Originais'!{letra_contrato}:{letra_contrato})-1"),
        ("V.B.:", f"=SUM('Dados Originais'!{letra_vb}:{letra_vb})"),
        ("V.L.:", f"=SUM('Dados Originais'!{letra_vl}:{letra_vl})"),
        ("IOF:", f"=SUM('Dados Originais'!{letra_iof}:{letra_iof})"),
        ("CAD:", f"=SUM('Dados Originais'!{letra_cad}:{letra_cad})"),
        ("Ágio:", "=ROUND(C4*0.0025,6)"),
        ("Imposto CAD:", "=ROUND(C7*0.0965,6)"),
        ("Valor Cessão:", "=C5+C6+C8+C9")
    ]

    _estilizar_bloco_resumo(ws, 2, "Resumo Sistema", "A9D18E", linhas)


def adicionar_resumo_parceiro_parceiro_b(wb):
    if "Resumo" not in wb.sheetnames:
        raise Exception("A aba 'Resumo' não foi encontrada no arquivo de resumo.")
    if "Dados Parceiro" not in wb.sheetnames:
        raise Exception("A aba 'Dados Parceiro' não foi encontrada no arquivo de resumo.")

    ws = wb["Resumo"]
    ws_parceiro = wb["Dados Parceiro"]

    col_operacao = encontrar_indice_coluna_por_nomes(ws_parceiro, ["Partner_Loan_ID", "Loan_ID", "CONTRATO"]) or 1
    col_vb = encontrar_indice_coluna_por_nomes(ws_parceiro, ["Valor CCB", "VALOR CCB", "Valor_CCB"])
    col_vl = encontrar_indice_coluna_por_nomes(ws_parceiro, ["Principal", "VALOR_LIQUIDO"])
    col_iof = encontrar_indice_coluna_por_nomes(ws_parceiro, ["IOF"])
    col_cad = encontrar_indice_coluna_por_nomes(ws_parceiro, ["TAC", "CAD"])

    if not all([col_vb, col_vl, col_iof, col_cad]):
        raise Exception("Não encontrei as colunas Valor CCB, Principal, IOF e TAC em 'Dados Parceiro'.")

    letra_op = get_column_letter(col_operacao)
    letra_vb = get_column_letter(col_vb)
    letra_vl = get_column_letter(col_vl)
    letra_iof = get_column_letter(col_iof)
    letra_cad = get_column_letter(col_cad)

    linhas = [
        ("Op.:", f"=COUNTA('Dados Parceiro'!{letra_op}:{letra_op})-1"),
        ("V.B.:", f"=SUM('Dados Parceiro'!{letra_vb}:{letra_vb})"),
        ("V.L.:", f"=SUM('Dados Parceiro'!{letra_vl}:{letra_vl})"),
        ("IOF:", f"=SUM('Dados Parceiro'!{letra_iof}:{letra_iof})"),
        ("CAD:", f"=SUM('Dados Parceiro'!{letra_cad}:{letra_cad})")
    ]

    _estilizar_bloco_resumo(ws, 12, "Resumo Parceiro", "00B0F0", linhas)




def normalizar_contrato(valor):
    if pd.isna(valor):
        return ""
    valor = str(valor).upper().strip()
    valor = valor.replace("\n", "")
    valor = valor.replace("\r", "")
    valor = valor.replace("\t", "")
    valor = valor.replace(" ", "")
    valor = valor.replace("–", "-").replace("—", "-")
    return valor


def normalizar_cpf(valor):
    if pd.isna(valor):
        return ""
    return (
        str(valor)
        .strip()
        .replace(".", "")
        .replace("-", "")
        .replace("/", "")
        .replace(" ", "")
    )


def extrair_base_contrato_parceiro_b(valor):
    """Retorna apenas a base de 8 caracteres do contrato/CCB (ex.: A1234567)."""
    texto = normalizar_contrato(valor)
    if not texto:
        return ""

    texto_sem_separadores = re.sub(r"[^A-Z0-9]", "", texto)
    match = re.search(r"A\d{7}", texto_sem_separadores)
    if match:
        return match.group(0)

    return texto_sem_separadores[:8]


def extrair_dados_pdf_endosso_parceiro_b(caminho_pdf):
    linhas = []
    contratos_vistos = set()

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            tabelas = pagina.extract_tables() or []
            for tabela in tabelas:
                for linha in tabela:
                    if not linha:
                        continue

                    primeira_coluna = str(linha[0] or "")
                    primeira_coluna = primeira_coluna.replace("\n", "")
                    primeira_coluna = primeira_coluna.replace("\r", "")
                    primeira_coluna = primeira_coluna.replace("\t", "")
                    primeira_coluna = primeira_coluna.replace(" ", "")
                    primeira_coluna = primeira_coluna.replace("–", "-").replace("—", "-")
                    primeira_coluna = primeira_coluna.strip().upper()

                    if not primeira_coluna:
                        continue

                    if "NºDACCB" in primeira_coluna or "ANEXOAOTERMODEENDOSSO" in primeira_coluna:
                        continue

                    if re.fullmatch(r"-?0+", primeira_coluna):
                        continue

                    contrato = extrair_base_contrato_parceiro_b(primeira_coluna)
                    if not contrato or contrato in contratos_vistos:
                        continue

                    cpf = normalizar_cpf(linha[1] if len(linha) > 1 else "")
                    data_emissao = str(linha[2]).strip().replace("\n", " ") if len(linha) > 2 and linha[2] else ""
                    qtd_parcelas = str(linha[3]).strip().replace("\n", " ") if len(linha) > 3 and linha[3] else ""
                    vencimento = str(linha[4]).strip().replace("\n", " ") if len(linha) > 4 and linha[4] else ""
                    valor_emissao = limpar_valor_monetario(linha[5]) if len(linha) > 5 else 0.0
                    preco_aquisicao = limpar_valor_monetario(linha[6]) if len(linha) > 6 else 0.0

                    linhas.append([
                        contrato,
                        cpf,
                        data_emissao,
                        qtd_parcelas,
                        vencimento,
                        valor_emissao,
                        preco_aquisicao
                    ])
                    contratos_vistos.add(contrato)

        if not linhas:
            texto_completo = ""
            for pagina in pdf.pages:
                texto = pagina.extract_text() or ""
                texto_completo += texto + "\n"

            texto_completo = texto_completo.replace("-\n", "-")
            texto_completo = texto_completo.replace("\n", "")
            texto_completo = texto_completo.replace("\r", "")
            texto_completo = texto_completo.replace("\t", "")
            texto_completo = texto_completo.replace(" ", "")
            texto_completo = texto_completo.replace("–", "-").replace("—", "-")
            encontrados = re.findall(r"A\d{7}(?:-\d{3})?", texto_completo)

            for contrato_encontrado in encontrados:
                contrato = extrair_base_contrato_parceiro_b(contrato_encontrado)
                if contrato and contrato not in contratos_vistos:
                    linhas.append([contrato, "", "", "", "", 0.0, 0.0])
                    contratos_vistos.add(contrato)

    if not linhas:
        raise Exception("Não foi possível localizar os contratos no PDF do Termo de Endosso.")

    return linhas


def criar_aba_endosso_parceiro_b(wb, linhas_endosso):
    if "Endosso" in wb.sheetnames:
        del wb["Endosso"]

    ws = wb.create_sheet(title="Endosso")
    cabecalho = [
        "Nº da CCB",
        "CPF",
        "Data de emissão da CCB",
        "Quantidade total parcelas da CCB",
        "Data de vencimento da CCB",
        "Valor de emissão da CCB (R$)",
        "Preço de Aquisição (R$)"
    ]
    ws.append(cabecalho)

    for linha in linhas_endosso:
        ws.append(linha)

    aplicar_formatacao_padrao(ws, ["Valor de emissão da CCB (R$)", "Preço de Aquisição (R$)"])


def criar_aba_divergencia_endosso_parceiro_b(wb):
    nome_aba = "Divergência Endosso"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws = wb.create_sheet(title=nome_aba)
    ws.append(["CONTRATO", "ORIGEM"])

    ws_orig = wb["Dados Originais"]
    ws_end = wb["Endosso"]

    col_contrato_orig = encontrar_indice_coluna_por_nomes(ws_orig, ["CONTRATO", "Nº DA CCB", "NUM_CONTRATO", "CCB"])
    col_contrato_end = encontrar_indice_coluna_por_nomes(ws_end, ["Nº da CCB", "CONTRATO", "CCB"])

    if col_contrato_orig is None or col_contrato_end is None:
        raise Exception("Não encontrei a coluna de contrato para gerar a aba de divergência do endosso.")

    contratos_orig = set()
    for row in range(2, ws_orig.max_row + 1):
        valor = extrair_base_contrato_parceiro_b(ws_orig.cell(row=row, column=col_contrato_orig).value)
        if valor:
            contratos_orig.add(valor)

    contratos_end = set()
    for row in range(2, ws_end.max_row + 1):
        valor = extrair_base_contrato_parceiro_b(ws_end.cell(row=row, column=col_contrato_end).value)
        if valor:
            contratos_end.add(valor)

    faltando_no_endosso = sorted(contratos_orig - contratos_end)
    faltando_no_original = sorted(contratos_end - contratos_orig)

    for contrato in faltando_no_endosso:
        ws.append([contrato, "Está em Dados Originais e não está no Endosso"])

    for contrato in faltando_no_original:
        ws.append([contrato, "Está no Endosso e não está em Dados Originais"])

    aplicar_formatacao_padrao(ws, [])
    return len(faltando_no_endosso) + len(faltando_no_original)


def localizar_linha_rotulo(ws, rotulo, ocorrencia=1, coluna=2):
    contador = 0
    rotulo = str(rotulo).strip().upper()
    for row in range(1, ws.max_row + 1):
        valor = ws.cell(row=row, column=coluna).value
        if str(valor).strip().upper() == rotulo:
            contador += 1
            if contador == ocorrencia:
                return row
    return None


def adicionar_tabela_checagem_endosso_parceiro_b(wb):
    if "Resumo" not in wb.sheetnames:
        raise Exception("A aba 'Resumo' não foi encontrada.")
    if "Dados Originais" not in wb.sheetnames:
        raise Exception("A aba 'Dados Originais' não foi encontrada.")
    if "Dados Parceiro" not in wb.sheetnames:
        raise Exception("A aba 'Dados Parceiro' não foi encontrada. Faça primeiro a conferência Excel x CSV.")
    if "Endosso" not in wb.sheetnames:
        raise Exception("A aba 'Endosso' não foi encontrada.")

    ws = wb["Resumo"]
    ws_orig = wb["Dados Originais"]
    ws_parc = wb["Dados Parceiro"]
    ws_end = wb["Endosso"]

    col_cpf_orig = encontrar_indice_coluna_por_nomes(ws_orig, ["CPF", "tax_id"])
    col_contrato_orig = encontrar_indice_coluna_por_nomes(ws_orig, ["CONTRATO", "Nº DA CCB", "NUM_CONTRATO", "CCB"])
    col_cpf_parc = encontrar_indice_coluna_por_nomes(ws_parc, ["tax_id", "CPF", "CPF/CNPJ"])
    col_contrato_parc = encontrar_indice_coluna_por_nomes(ws_parc, ["Partner_Loan_ID", "Loan_ID", "CONTRATO", "CCB"])
    col_contrato_end = encontrar_indice_coluna_por_nomes(ws_end, ["Nº da CCB", "CONTRATO", "CCB"])

    if None in [col_cpf_orig, col_contrato_orig, col_cpf_parc, col_contrato_parc, col_contrato_end]:
        raise Exception("Não encontrei uma ou mais colunas obrigatórias para a checagem do PARCEIRO B Endosso.")

    def coletar_valores(ws_origem, coluna, normalizador=str):
        valores = []
        for row in range(2, ws_origem.max_row + 1):
            valor = ws_origem.cell(row=row, column=coluna).value
            valor = normalizador(valor)
            if valor and valor.upper() not in {"CPF", "TAX_ID", "CONTRATO", "CCB", "PARTNER_LOAN_ID", "LOAN_ID", "Nº DA CCB"}:
                valores.append(valor)
        return valores

    cpfs_orig = coletar_valores(ws_orig, col_cpf_orig, normalizar_cpf)
    cpfs_parc = coletar_valores(ws_parc, col_cpf_parc, normalizar_cpf)
    contratos_orig = coletar_valores(ws_orig, col_contrato_orig, extrair_base_contrato_parceiro_b)
    contratos_parc = coletar_valores(ws_parc, col_contrato_parc, extrair_base_contrato_parceiro_b)
    contratos_end = coletar_valores(ws_end, col_contrato_end, extrair_base_contrato_parceiro_b)

    status_cpf = "CONFERE" if sorted(cpfs_orig) == sorted(cpfs_parc) else "NÃO CONFERE"
    status_ccb = "CONFERE" if sorted(contratos_orig) == sorted(contratos_parc) else "NÃO CONFERE"
    status_endosso = "CONFERE" if sorted(contratos_orig) == sorted(contratos_end) else "NÃO CONFERE"

    linha_op_sistema = localizar_linha_rotulo(ws, "Op.:", 1)
    linha_vb_sistema = localizar_linha_rotulo(ws, "V.B.:", 1)
    linha_vl_sistema = localizar_linha_rotulo(ws, "V.L.:", 1)
    linha_iof_sistema = localizar_linha_rotulo(ws, "IOF:", 1)
    linha_cad_sistema = localizar_linha_rotulo(ws, "CAD:", 1)

    linha_op_parceiro = localizar_linha_rotulo(ws, "Op.:", 2)
    linha_vb_parceiro = localizar_linha_rotulo(ws, "V.B.:", 2)
    linha_vl_parceiro = localizar_linha_rotulo(ws, "V.L.:", 2)
    linha_iof_parceiro = localizar_linha_rotulo(ws, "IOF:", 2)
    linha_cad_parceiro = localizar_linha_rotulo(ws, "CAD:", 2)

    if None in [linha_op_sistema, linha_vb_sistema, linha_vl_sistema, linha_iof_sistema, linha_cad_sistema,
                linha_op_parceiro, linha_vb_parceiro, linha_vl_parceiro, linha_iof_parceiro, linha_cad_parceiro]:
        raise Exception("Não consegui localizar as linhas do Resumo Sistema e do Resumo Parceiro.")

    def status_por_formula(cel_sistema, cel_parceiro):
        return f'=IF(ROUND({cel_sistema},2)=ROUND({cel_parceiro},2),"CONFERE","NÃO CONFERE")'

    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    fonte_texto = Font(name="Aptos Narrow", size=11, bold=True, color="000000")

    def escrever_bloco(linha_inicial, titulo, itens, cor_titulo):
        ws.merge_cells(start_row=linha_inicial, start_column=8, end_row=linha_inicial, end_column=9)
        c = ws.cell(row=linha_inicial, column=8, value=titulo)
        c.font = fonte_titulo
        c.fill = PatternFill(start_color=cor_titulo, end_color=cor_titulo, fill_type="solid")
        c.alignment = alinhamento
        ws.cell(row=linha_inicial, column=8).border = borda
        ws.cell(row=linha_inicial, column=9).border = borda

        for i, (rotulo, valor) in enumerate(itens, start=1):
            r = linha_inicial + i
            c1 = ws.cell(row=r, column=8, value=rotulo)
            c2 = ws.cell(row=r, column=9, value=valor)
            c1.font = fonte_texto
            c2.font = fonte_texto
            c1.alignment = alinhamento
            c2.alignment = alinhamento
            c1.border = borda
            c2.border = borda

    itens_checagem = [
        ("CPF", status_cpf),
        ("CCB", status_ccb),
        ("Op.:", f'=IF(C{linha_op_sistema}=C{linha_op_parceiro},"CONFERE","NÃO CONFERE")'),
        ("V.B.", status_por_formula(f"C{linha_vb_sistema}", f"C{linha_vb_parceiro}")),
        ("V.L.", status_por_formula(f"C{linha_vl_sistema}", f"C{linha_vl_parceiro}")),
        ("IOF", status_por_formula(f"C{linha_iof_sistema}", f"C{linha_iof_parceiro}")),
        ("CAD", status_por_formula(f"C{linha_cad_sistema}", f"C{linha_cad_parceiro}"))
    ]

    itens_endosso = [("CCB", status_endosso)]

    escrever_bloco(2, "CHECAGEM", itens_checagem, "FFF200")
    escrever_bloco(11, "TERMO DE ENDOSSO", itens_endosso, "FF0000")

    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18

    verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    verde_font = Font(color="008000")
    vermelho_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    vermelho_font = Font(color="9C0006")

    for faixa in ["I3:I9", "I12:I12"]:
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(operator="equal", formula=['"CONFERE"'], fill=verde_fill, font=verde_font)
        )
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(operator="equal", formula=['"NÃO CONFERE"'], fill=vermelho_fill, font=vermelho_font)
        )


def fazer_conferencia_parceiro_b_endosso():
    messagebox.showinfo(
        "Conferência PARCEIRO B - Endosso",
        "1º Selecione o arquivo EXCEL de Resumo\n2º Selecione o PDF do Termo de Endosso"
    )

    caminho_excel = filedialog.askopenfilename(
        title="Selecionar Resumo Excel",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_excel:
        return

    caminho_pdf = filedialog.askopenfilename(
        title="Selecionar PDF Termo de Endosso",
        filetypes=[("PDF", "*.pdf")]
    )
    if not caminho_pdf:
        return

    try:
        wb = load_workbook(caminho_excel)
        linhas_endosso = extrair_dados_pdf_endosso_parceiro_b(caminho_pdf)

        criar_aba_endosso_parceiro_b(wb, linhas_endosso)
        qtd_divergencias = criar_aba_divergencia_endosso_parceiro_b(wb)
        adicionar_tabela_checagem_endosso_parceiro_b(wb)

        try:
            wb.save(caminho_excel)
        except PermissionError:
            messagebox.showerror("Arquivo aberto", f"Feche o arquivo no Excel e tente novamente:\n{caminho_excel}")
            return

        msg = "Conferência do Endosso incluída com sucesso!"
        if qtd_divergencias:
            msg += f"\nForam encontradas {qtd_divergencias} divergências na aba 'Divergência Endosso'."

        messagebox.showinfo("Sucesso", msg)
        os.startfile(caminho_excel)

    except Exception as erro:
        messagebox.showerror("Erro na Conferência PARCEIRO B - Endosso", str(erro))

def selecionar_e_processar_parceiro_b(root, entry_data, progress_bar, nome_parceiro, codigo_parceiro, event=None):
    data_digitada = entry_data.get().strip()

    if not validar_data(data_digitada):
        messagebox.showwarning("Atenção", "Por favor, digite uma data válida no formato dd.mm.aaaa.")
        return

    caminho_arquivo = filedialog.askopenfilename(
        title=f"Selecionar CSV do Sistema - {nome_parceiro}",
        filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        progress_bar.pack(pady=10)
        progress_bar.set(0.25)
        root.update_idletasks()

        df = carregar_arquivo(caminho_arquivo)
        df = padronizar_colunas(df)
        df = converter_colunas_numericas(df)

        progress_bar.set(0.60)
        root.update_idletasks()

        pasta_saida = os.path.dirname(caminho_arquivo)
        nome_curto = obter_nome_curto_parceiro_b(codigo_parceiro, nome_parceiro)

        nome_arquivo_resumo = obter_nome_resumo_parceiro_b(codigo_parceiro, data_digitada, nome_parceiro)
        caminho_resumo = os.path.join(pasta_saida, nome_arquivo_resumo)

        nome_arquivo_exportacao = obter_nome_exportacao_parceiro_b(codigo_parceiro, data_digitada, nome_parceiro)
        caminho_exportacao = os.path.join(pasta_saida, nome_arquivo_exportacao)

        wb = Workbook()
        criar_aba_dados_originais(wb, df)
        criar_aba_resumo_parceiro_b_sistema(wb)
        wb.save(caminho_resumo)

        salvar_planilha_parceiro_b(df, caminho_exportacao)

        progress_bar.set(1.0)
        root.update_idletasks()
        progress_bar.pack_forget()

        os.startfile(pasta_saida)
        messagebox.showinfo("Sucesso", f"Arquivos gerados com sucesso para {nome_curto}!")

    except Exception as erro:
        progress_bar.pack_forget()
        messagebox.showerror("Erro", str(erro))


def fazer_conferencia_parceiro_b():
    messagebox.showinfo(
        "Conferência PARCEIRO B",
        "1º Selecione o arquivo EXCEL de Resumo\n2º Selecione o CSV 'RELATÓRIO' do parceiro"
    )

    caminho_excel = filedialog.askopenfilename(
        title="Selecionar Resumo Excel",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_excel:
        return

    caminho_csv = filedialog.askopenfilename(
        title="Selecionar CSV RELATÓRIO PARCEIRO B",
        filetypes=[("CSV", "*.csv")]
    )
    if not caminho_csv:
        return

    try:
        wb = load_workbook(caminho_excel)
        df_parceiro = carregar_arquivo(caminho_csv)
        df_parceiro = padronizar_colunas(df_parceiro)

        criar_aba_dados_parceiro_parceiro_b(wb, df_parceiro)
        adicionar_resumo_parceiro_parceiro_b(wb)

        try:
            wb.save(caminho_excel)
        except PermissionError:
            messagebox.showerror("Arquivo aberto", f"Feche o arquivo no Excel e tente novamente:\n{caminho_excel}")
            return

        messagebox.showinfo("Sucesso", "Resumo do parceiro PARCEIRO B incluído com sucesso!")
        os.startfile(caminho_excel)

    except Exception as erro:
        messagebox.showerror("Erro na Conferência PARCEIRO B", str(erro))