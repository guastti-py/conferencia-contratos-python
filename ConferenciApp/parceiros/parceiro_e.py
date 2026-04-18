# ===================================================================
# PARCEIRO: PARCEIRO E
# Regras de conferência entre PDF do termo e Excel de resumo
# ===================================================================

import os
import pandas as pd
import pdfplumber
from tkinter import messagebox, filedialog

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from utils.arquivos import limpar_valor_monetario
from utils.excel import aplicar_formatacao_padrao


def encontrar_coluna_contrato(df):
    for coluna in df.columns:
        nome_coluna = str(coluna).upper()
        if "CONTRATO" in nome_coluna or "CCB" in nome_coluna:
            return coluna
    return df.columns[0]


def extrair_dados_pdf_parceiro_e(caminhos_pdf):
    """
    Extrai do(s) PDF(s) do Termo PARCEIRO E.
    Colunas: col0=Nº CCB | col1=Emitente | col2=CPF | col3=Valor bruto
    """
    todos_ccbs = []
    todos_cpfs = []
    total_valor_bruto = 0.0
    linhas_aba = []

    for caminho_pdf in caminhos_pdf:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                for tabela in pagina.extract_tables():
                    for linha in tabela:
                        if not linha or not linha[0]:
                            continue
                        col0 = str(linha[0]).strip().replace("\n", " ")
                        if "Nº da CCB" in col0 or col0.startswith("Total") or col0 == "":
                            continue
                        ccb  = col0
                        nome = str(linha[1]).strip().replace("\n", " ") if len(linha) > 1 and linha[1] else ""
                        cpf  = str(linha[2]).strip().replace("\n", " ").replace(".", "").replace("-", "") if len(linha) > 2 and linha[2] else ""
                        vb   = limpar_valor_monetario(linha[3]) if len(linha) > 3 and linha[3] else 0.0

                        todos_ccbs.append(ccb)
                        todos_cpfs.append(cpf)
                        total_valor_bruto += vb
                        linhas_aba.append([ccb, nome, cpf, vb])

    return {
        "ccbs": set(todos_ccbs),
        "cpfs": set(todos_cpfs),
        "total_valor_bruto": round(total_valor_bruto, 2),
        "total_contratos": len(todos_ccbs),
        "linhas_aba": linhas_aba
    }


def salvar_aba_dados_pdf_parceiro_e(wb, linhas_aba):
    """Cria (ou recria) a aba 'Dados PDF' com as linhas extraídas do(s) Termo(s) PARCEIRO E."""
    nome_aba = "Dados PDF"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws_pdf = wb.create_sheet(title=nome_aba)
    cabecalho = ["Nº da CCB", "Emitente", "CPF do Emitente", "Valor Bruto"]
    ws_pdf.append(cabecalho)

    for linha in linhas_aba:
        ws_pdf.append(linha)

    ultima_linha = ws_pdf.max_row + 1
    ws_pdf.cell(row=ultima_linha, column=1, value="TOTAL")
    ws_pdf.cell(row=ultima_linha, column=4, value=round(sum(l[3] for l in linhas_aba), 2))

    aplicar_formatacao_padrao(ws_pdf, ["Valor Bruto"])

    fill_total = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fonte_total = Font(bold=True)
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    alin = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        c = ws_pdf.cell(row=ultima_linha, column=col)
        c.fill = fill_total
        c.font = fonte_total
        c.border = borda
        c.alignment = alin
        if col == 4:
            c.number_format = 'R$ #,##0.00'


def obter_dados_resumo_sistema_parceiro_e(caminho_excel):
    df = pd.read_excel(caminho_excel, sheet_name="Dados Originais")
    df.columns = df.columns.str.strip()

    col_ccb = None
    for col in df.columns:
        if "CONTRATO" in col.upper() or "CCB" in col.upper():
            col_ccb = col
            break

    col_cpf = None
    for col in df.columns:
        if "CPF" in col.upper():
            col_cpf = col
            break

    col_vl = None
    for col in df.columns:
        if "VALOR_LIQUIDO" in col.upper() or "VALOR LIQUIDO" in col.upper():
            col_vl = col
            break

    ccbs = set(df[col_ccb].astype(str).str.strip().tolist()) if col_ccb else set()

    cpfs = set()
    if col_cpf:
        cpfs = set(
            df[col_cpf].astype(str).str.strip()
            .str.replace(".", "", regex=False)
            .str.replace("-", "", regex=False)
            .tolist()
        )

    vl_total = 0.0
    if col_vl:
        vl_total = pd.to_numeric(df[col_vl], errors="coerce").fillna(0).sum()

    return {
        "ccbs": ccbs,
        "cpfs": cpfs,
        "total_contratos": len(df),
        "total_vl": round(vl_total, 2)
    }


def escrever_resumo_termo_parceiro_e(ws, linha_inicio, col_inicio, dados_pdf):
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(name="Aptos Narrow", size=13, bold=True, color="FFFFFF")
    fonte_valor = Font(name="Aptos Narrow", size=12, bold=True)
    cor_azul = "2E75B6"

    ws.merge_cells(
        start_row=linha_inicio, start_column=col_inicio,
        end_row=linha_inicio, end_column=col_inicio + 1
    )
    cel = ws.cell(row=linha_inicio, column=col_inicio, value="Resumo Termo")
    cel.font = fonte_titulo
    cel.fill = PatternFill(start_color=cor_azul, end_color=cor_azul, fill_type="solid")
    cel.alignment = alinhamento
    ws.cell(row=linha_inicio, column=col_inicio).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa)
    ws.cell(row=linha_inicio, column=col_inicio + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa)

    for i, (rotulo, valor, fmt) in enumerate([
        ("Op.:",     dados_pdf["total_contratos"],  False),
        ("V.Bruto:", dados_pdf["total_valor_bruto"], True),
    ], start=1):
        eh_ultima = i == 2
        cel_r = ws.cell(row=linha_inicio + i, column=col_inicio, value=rotulo)
        cel_r.font = fonte_valor
        cel_r.alignment = alinhamento
        cel_r.border = Border(
            left=borda_grossa, right=borda_fina,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina)

        cel_v = ws.cell(row=linha_inicio + i, column=col_inicio + 1, value=valor)
        cel_v.font = fonte_valor
        cel_v.alignment = alinhamento
        cel_v.border = Border(
            left=borda_fina, right=borda_grossa,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina)
        if fmt:
            cel_v.number_format = 'R$ #,##0.00'


def escrever_checagem_parceiro_e(ws, linha_inicio, col_inicio, dados_sistema, dados_pdf):
    borda_fina = Side(style="thin", color="000000")
    borda_grossa = Side(style="medium", color="000000")
    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte_titulo = Font(name="Aptos Narrow", size=13, bold=True, color="000000")
    fonte_header = Font(name="Aptos Narrow", size=11, bold=True)
    fonte_item   = Font(name="Aptos Narrow", size=11)
    verde_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    verde_font   = Font(color="008000", bold=True)
    vermelho_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    vermelho_font = Font(color="9C0006", bold=True)

    ws.merge_cells(
        start_row=linha_inicio, start_column=col_inicio,
        end_row=linha_inicio, end_column=col_inicio + 1
    )
    cel = ws.cell(row=linha_inicio, column=col_inicio, value="CHECAGEM")
    cel.font = fonte_titulo
    cel.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cel.alignment = alinhamento
    ws.cell(row=linha_inicio, column=col_inicio).border = Border(
        left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa)
    ws.cell(row=linha_inicio, column=col_inicio + 1).border = Border(
        left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa)

    cel_p = ws.cell(row=linha_inicio + 1, column=col_inicio, value="PRODUTO")
    cel_p.font = fonte_header
    cel_p.alignment = alinhamento
    cel_p.border = Border(left=borda_grossa, right=borda_fina, top=borda_grossa, bottom=borda_grossa)
    cel_s = ws.cell(row=linha_inicio + 1, column=col_inicio + 1, value="STATUS")
    cel_s.font = fonte_header
    cel_s.alignment = alinhamento
    cel_s.border = Border(left=borda_fina, right=borda_grossa, top=borda_grossa, bottom=borda_grossa)

    itens = [
        ("CPF",     dados_sistema["cpfs"]            == dados_pdf["cpfs"]),
        ("CCB",     dados_sistema["ccbs"]            == dados_pdf["ccbs"]),
        ("Nº Op.",  dados_sistema["total_contratos"] == dados_pdf["total_contratos"]),
        ("V.BRUTO", round(dados_sistema["total_vl"], 2) == round(dados_pdf["total_valor_bruto"], 2)),
    ]

    for i, (produto, ok) in enumerate(itens):
        linha = linha_inicio + 2 + i
        eh_ultima = i == len(itens) - 1
        status = "CONFERE" if ok else "NÃO CONFERE"

        cel_pr = ws.cell(row=linha, column=col_inicio, value=produto)
        cel_pr.font = fonte_item
        cel_pr.alignment = alinhamento
        cel_pr.border = Border(
            left=borda_grossa, right=borda_fina,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina)

        cel_st = ws.cell(row=linha, column=col_inicio + 1, value=status)
        cel_st.alignment = alinhamento
        cel_st.border = Border(
            left=borda_fina, right=borda_grossa,
            top=borda_fina, bottom=borda_grossa if eh_ultima else borda_fina)
        if ok:
            cel_st.fill = verde_fill
            cel_st.font = verde_font
        else:
            cel_st.fill = vermelho_fill
            cel_st.font = vermelho_font


def fazer_conferencia():
    messagebox.showinfo(
        "Conferência PARCEIRO E",
        "1º Selecione o arquivo EXCEL de Resumo\n"
        "2º Selecione o(s) PDF(s) do Termo de Endosso"
    )

    caminho_excel = filedialog.askopenfilename(
        title="Selecionar Resumo Excel - PARCEIRO E",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not caminho_excel:
        return

    caminhos_pdf = []
    while True:
        caminho_pdf = filedialog.askopenfilename(
            title=f"Selecionar PDF do Termo ({len(caminhos_pdf) + 1}º)",
            filetypes=[("PDF", "*.pdf")]
        )
        if not caminho_pdf:
            if not caminhos_pdf:
                messagebox.showwarning("Atenção", "Nenhum PDF selecionado. Operação cancelada.")
                return
            break
        caminhos_pdf.append(caminho_pdf)
        resposta = messagebox.askyesno(
            "Mais PDFs?",
            f"{len(caminhos_pdf)} PDF(s) selecionado(s).\n\nDeseja adicionar mais um PDF?"
        )
        if not resposta:
            break

    try:
        dados_pdf     = extrair_dados_pdf_parceiro_e(caminhos_pdf)
        dados_sistema = obter_dados_resumo_sistema_parceiro_e(caminho_excel)

        wb = load_workbook(caminho_excel)

        if "Resumo" not in wb.sheetnames:
            raise Exception("A aba 'Resumo' não foi encontrada no Excel.")

        salvar_aba_dados_pdf_parceiro_e(wb, dados_pdf["linhas_aba"])

        ws = wb["Resumo"]

        ultima_linha_usada = 1
        for row in ws.iter_rows(min_col=2, max_col=3):
            for cell in row:
                if cell.value is not None:
                    ultima_linha_usada = max(ultima_linha_usada, cell.row)

        linha_resumo_termo = ultima_linha_usada + 2
        linha_checagem = 2

        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 18

        escrever_resumo_termo_parceiro_e(ws, linha_resumo_termo, 2, dados_pdf)
        escrever_checagem_parceiro_e(ws, linha_checagem, 5, dados_sistema, dados_pdf)

        try:
            wb.save(caminho_excel)
        except PermissionError:
            messagebox.showerror(
                "Arquivo aberto",
                f"Feche o arquivo no Excel e tente novamente:\n{caminho_excel}"
            )
            return

        messagebox.showinfo(
            "Sucesso",
            f"Conferência PARCEIRO E concluída!\n"
            f"{len(caminhos_pdf)} PDF(s) processado(s).\n"
            f"Total contratos PDF: {dados_pdf['total_contratos']}\n"
            f"V.Bruto PDF: R$ {dados_pdf['total_valor_bruto']:,.2f}"
        )
        os.startfile(caminho_excel)

    except Exception as erro:
        messagebox.showerror("Erro na Conferência PARCEIRO E", str(erro))
